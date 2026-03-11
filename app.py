import os
import json
import uuid
import base64
from datetime import datetime, timedelta
from typing import Optional

import resend
import stripe
from apscheduler.schedulers.background import BackgroundScheduler
from flask import Flask, request, jsonify, send_file, render_template, redirect
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

from emails import email_1_html, email_2_html, email_3_html, email_4_html, email_5_html, email_abandoned_html

load_dotenv()
stripe.api_key = os.environ.get('STRIPE_SECRET_KEY')
resend.api_key = os.environ.get('RESEND_API_KEY')
FROM_EMAIL = os.environ.get('FROM_EMAIL', 'noreply@example.com')

app = Flask(__name__)
client = OpenAI(api_key=os.environ.get("OPENAI_API_KEY"))

_PLAN_CACHE: dict = {}
_UNSUB_TOKENS: dict = {}

scheduler = BackgroundScheduler()
scheduler.start()


# ─────────────────────────────────────────────────────────────────────────────
# EMAIL
# ─────────────────────────────────────────────────────────────────────────────

def send_email(to: str, subject: str, html: str, attachment_path: Optional[str] = None):
    params: dict = {"from": FROM_EMAIL, "to": [to], "subject": subject, "html": html}
    if attachment_path and os.path.exists(attachment_path):
        with open(attachment_path, "rb") as f:
            content = base64.b64encode(f.read()).decode()
        params["attachments"] = [{"filename": os.path.basename(attachment_path), "content": content}]
    try:
        r = resend.Emails.send(params)
        app.logger.info(f"Resend OK: {r.get('id')} → {to}")
    except Exception as exc:
        app.logger.error(f"Resend FAILED for {to}: {exc}")


def schedule_email_sequence(to_email: str, name: str, file_id: str, host_url: str):
    free_url    = f"{host_url}download/free/{file_id}"
    payment_url = f"{host_url}payment?file_id={file_id}"
    free_path   = f"/tmp/free_{file_id}.xlsx"
    now         = datetime.now()

    unsub_token = str(uuid.uuid4())
    _UNSUB_TOKENS[unsub_token] = {"email": to_email, "file_id": file_id}
    unsubscribe_url = f"{host_url}unsubscribe/{unsub_token}"

    sequence = [
        (1, "The #1 thing coaches who fill their calendar do differently",
         email_2_html(name, free_url, payment_url, unsubscribe_url), False),
        (2, "What most coaches skip — and why it quietly kills their revenue",
         email_3_html(name, free_url, payment_url, unsubscribe_url), False),
        (4, f"{name.split()[0] if name else 'Coach'}, here's what stood out in your quiz answers",
         email_4_html(name, free_url, payment_url, unsubscribe_url), False),
        (6, "This is the last time I'll mention it",
         email_5_html(name, free_url, payment_url, unsubscribe_url), False),
    ]
    for delay_days, subject, html, attach in sequence:
        run_at     = now + timedelta(days=delay_days)
        attachment = free_path if attach else None
        scheduler.add_job(
            send_email, trigger="date", run_date=run_at,
            args=[to_email, subject, html, attachment],
            id=f"{file_id}_email_{delay_days}", replace_existing=True,
        )
    app.logger.info(f"Drip scheduled for {to_email}")


# ─────────────────────────────────────────────────────────────────────────────
# PATH DETECTION — FOUR PATHS
#
#  A → Has a group program, running paid ads to fill it, established coach.
#      The cohort resets every cycle. The problem is a leaky acquisition
#      engine and no continuity — NOT capacity. Fix: organic engine +
#      alumni continuity + cohort waitlist system.
#
#  B → Still building. No group program, beginner or early-stage, low/mid
#      pricing. Needs offer clarity, funnel, and first clients.
#
#  C → Established coach already doing corporate/B2B work (workshops,
#      organisational clients, corporate wellness) but NOT on speaking
#      stages yet. The next unlock is paid speaking and keynotes.
#
#  D → Established coach with premium somatic/wellness/performance niche,
#      NOT in corporate yet. Has untapped organisational demand they
#      haven't seen. Needs a corporate adoption strategy.
#
# DETECTION PHILOSOPHY — signal scoring, not rigid rules:
#   - No single field gates a path. Signals are combined and weighted.
#   - Duration is soft context, not a hard filter.
#   - When signals are mixed, the strongest cluster wins.
#   - Fallback is always B (never leave someone without a plan).
# ─────────────────────────────────────────────────────────────────────────────

def _detect_path(answers: dict) -> str:
    """
    Return 'A', 'B', 'C', or 'D' based on weighted quiz signals.
    """
    # ── raw fields ──────────────────────────────────────────────────────────
    fork         = (answers.get("QA_FORK") or "").lower()
    experience   = (answers.get("Q3")  or answers.get("QA3_exp") or "").lower()
    ideal_client = (answers.get("Q4")  or answers.get("QA4")     or "").lower()
    niche        = (answers.get("Q1")  or answers.get("QA1")     or "").lower()
    pain_point   = (answers.get("Q5")  or answers.get("QA5")     or "").lower()
    pricing      = (answers.get("Q10") or answers.get("QA10")    or "").lower()
    delivery     = (answers.get("Q8")  or answers.get("QA8")     or "").lower()
    duration     = (answers.get("Q9")  or answers.get("QA9")     or "").lower()
    channels     = (answers.get("Q17") or answers.get("QA7")     or "").lower()
    hours        = (answers.get("Q14") or answers.get("QA13")    or "").lower()
    paid_flag    = (answers.get("Q15_interest") or "").lower()

    # ── base signals ────────────────────────────────────────────────────────
    is_advanced = "advanced" in experience or "established" in experience
    is_beginner = "beginner" in experience or "just start" in experience

    # premium = pricing tier starts at $2k+; excludes "$500 - $2,000" (substring trap)
    premium_pricing = any(p in pricing for p in
        ["$2,000 - $5,000", "$5,000", "5,000+", "$2,000+", "2000 - 5000"])

    # group delivery: explicitly group/cohort/course — NOT hybrid (hybrid = 1:1 + digital)
    has_group_delivery = any(k in delivery for k in
        ["group", "cohort", "course + "])

    uses_paid_ads = (
        paid_flag == "yes" or
        "paid ads" in channels or
        "facebook" in channels or
        "paid" in channels
    )

    # corporate/speaking already active in channels
    already_speaking   = any(k in channels for k in
        ["speaking", "keynote", "stage", "bureau"])
    already_corporate  = any(k in channels for k in
        ["corporate", "b2b", "organisations", "companies",
         "workshop", "speaking/workshop"])

    # somatic / wellness / performance niche
    somatic_keywords = [
        "nervous system", "somatic", "fascial", "fascia", "regulation",
        "resilience", "burnout", "stress", "trauma", "performance",
        "leadership", "wellbeing", "well-being", "mindfulness", "breath",
        "embodied", "self-override", "friction", "flow", "body"
    ]
    has_somatic_niche = any(kw in niche + " " + pain_point
                            for kw in somatic_keywords)

    # corporate-adjacent ideal client
    corp_client_keywords = [
        "professional", "executive", "leader", "manager", "team", "corporate",
        "organisation", "organization", "company", "companies", "parent",
        "family", "families", "high-functioning", "high functioning",
        "performance", "employee", "workforce", "staff"
    ]
    has_corporate_client = any(kw in ideal_client for kw in corp_client_keywords)

    # hours maxed (soft signal that capacity is near limit)
    hours_high = "30" in hours or "full" in hours.lower()

    # ── PATH A — leaky group cohort coach ───────────────────────────────────
    # Hard gates: must be advanced + running a group program + using paid ads
    # Soft signals add confidence — duration is context, not a gate
    a_soft = sum([
        premium_pricing,                                       # charges real money
        hours_high,                                            # near capacity
        "4 week" in duration or "8 week" in duration,         # short cycle (soft)
        not already_corporate,                                 # hasn't expanded yet
    ])
    if is_advanced and has_group_delivery and uses_paid_ads and a_soft >= 1:
        return "A"

    # ── PATH C — corporate-active, no speaking yet ──────────────────────────
    # Core: already doing corporate/B2B work, speaking not yet monetised
    if is_advanced and already_corporate and not already_speaking:
        return "C"

    # ── PATH D — established somatic/premium, not in corporate yet ──────────
    # Core: advanced + premium + niche maps to corporate demand + no corp yet
    d_score = sum([
        is_advanced,
        premium_pricing,
        has_somatic_niche or has_corporate_client,
        not already_corporate,
        not uses_paid_ads,          # not running ads = not in Path A territory
    ])
    if d_score >= 4:
        return "D"

    # ── PATH B — fallback (beginner / building) ─────────────────────────────
    return "B"


# ─────────────────────────────────────────────────────────────────────────────
# OPENAI SYSTEM PROMPTS
# ─────────────────────────────────────────────────────────────────────────────

SYSTEM_PROMPT_B = """You are an expert coaching business strategist.

Your job: take quiz answers from a coach who is BUILDING their practice (Path B)
and return a complete personalised plan as structured JSON.

RULES:
- Return ONLY valid JSON. No markdown. No code blocks.
- Every field must be specific to THEIR answers — never generic.
- Use the client's own words from QA3 (client problem) throughout.
- Never use coaching platitudes like "do the inner work" or "step into your power".
- Offer sentence formula: I help [WHO] who [PROBLEM IN THEIR EXACT WORDS] to [TANGIBLE OUTCOME] in [TIMEFRAME] without [THEIR #1 OBJECTION]
- All revenue/funnel numeric fields MUST be plain integers — no $ signs, no text.
- For competitors: 3 REAL named people with actual pricing — never invent names.

MARKETING RULES:
- Every channel: profile audit, 3 pillars (3 hooks each), 3 complete sample posts (150-250 words), 10-entry 2-week calendar, 3 quick wins, 5 KPIs.
- sample_post must be publish-ready — write it AS the coach posting today.
- hooks: scroll-stopping first line. Start with pain, a number, or a counterintuitive statement. NEVER start with "I" or "Are you".
- two_week_calendar: exactly 10 entries Mon–Fri Wk1 then Mon–Fri Wk2. Vary formats.
- quick_wins: 3 most impactful actions in 48 hrs, fully written out, copy-paste ready.
- primary_metric: one number with a diagnostic rule attached.
- trust_channel_benchmarks used_now: "Yes" or "No" based on QA7 (channels).
- trust_channel_benchmarks estimated_annual_clients and estimated_annual_revenue: plain integers.

CHANNEL-SPECIFIC profile_audit keys:
  LinkedIn     → headline_before, headline_after, about_before, about_after, banner_tip, featured_section
  Instagram    → bio_before, bio_after, bio_formula, highlight_covers, story_strategy
  Email        → subject_line_formula, welcome_sequence, list_growth_tactic, segmentation_tip
  YouTube      → channel_description_rewrite, about_page_cta, thumbnail_formula, channel_trailer_script_outline
  Podcast      → show_description_rewrite, episode_title_formula, intro_hook_script, guest_pitch_template
  Facebook Ads → audience_definition, ad_hook_formula, creative_brief, landing_page_checklist
  Referrals    → referral_offer, ask_points, referral_script, tracking_system
  Partnerships → partner_target_list, partner_offer, outreach_script, co_marketing_asset"""


SYSTEM_PROMPT_A = """You are an expert coaching business strategist specialising in B2B revenue expansion.

Your job: take quiz answers from a coach who is FULLY BOOKED with 1:1 clients (Path A)
and return a plan that helps them scale revenue WITHOUT adding more 1:1 hours.

The two channels are: (1) Corporate Adoption Offer, (2) Speaking as a Revenue Channel.
These coaches do NOT need offer positioning or lead funnels — they already have clients.
They need to enter corporate doors that are already open with budget attached.

CRITICAL FRAMING RULES:
- Corporate buyers fund: change management, initiative adoption, retention programs, leadership transitions — NOT "coaching".
- The offer must enter through a budget line that already exists, not create a new one.
- Pricing anchors to COST AVOIDED (replacing one manager = $50–100K) not hourly rate.
- The speaking talk must NAME A PROBLEM the audience feels but can't articulate — not inspire.
- buying_trigger is as important as buyer_title: tell them WHEN to reach out, not just who.
- All numeric fields MUST be plain integers.
- Return ONLY valid JSON. No markdown. No code blocks."""


SYSTEM_PROMPT_C = """You are an expert B2B revenue strategist who helps established somatic, nervous system, and wellness coaches
monetise their IP through corporate adoption — not just information delivery.

Your job: analyse an established 1:1 coach and design a corporate ADOPTION strategy — not a speaking gig or an info product.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
RULE 1 — NUMBERS: ONLY USE WHAT THEY GAVE YOU
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
The quiz gives you a RANGE for pricing (e.g. "$2,000–$5,000") and a RANGE for hours (e.g. "15–30 hrs/week").
- For revenue ceiling calculations: use the MIDPOINT of each range and label it as an ESTIMATE.
- Always include the label "estimated from quiz range" next to calculated numbers.
- NEVER invent an exact hourly rate or exact weekly hours — derive from the range given.
- Example: pricing "$2,000–$5,000" → use $3,500 midpoint. Hours "15–30/week" → use 22hrs/week midpoint.
- ceiling_data_source must explain exactly which quiz answer was used and how the number was derived.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
RULE 2 — CORPORATE LANGUAGE: SPEAK IN HR BUDGET LINES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
HR directors, L&D managers, and Chief People Officers do NOT buy "coaching" or "wellness."
They buy solutions to problems that already have a budget line. Frame every offer in these terms:

WHAT HR ACTUALLY FUNDS:
- Burnout & absenteeism reduction (avg cost per absent employee: $3,000–$6,000/yr)
- Presenteeism programs (disengaged employee costs 34% of their salary)
- Retention initiatives (replacing one employee = 50–200% of their annual salary)
- Mental health & stress management (EAP alternative)
- Leadership resilience programs (L&D budget line)
- Return-to-work / parental transition programs
- Change management & nervous system safety during restructures

NEVER use: "coaching," "somatic work," "inner work," "embodiment," "fascial," "self-override"
ALWAYS use: "performance," "retention," "engagement," "resilience," "absenteeism," "presenteeism," "ROI"

The corporate_pain field must cite ONE of these budget lines with a real cost figure (e.g. "burnout costs UK employers £28bn/year").

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
RULE 3 — ADOPTION FUNNEL, NOT INFORMATION DELIVERY
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
The goal is NOT to educate HR. The goal is to get the methodology ADOPTED inside organisations.
Design a 3-stage adoption funnel:

STAGE 1 — FREE LEAD MAGNET (gets you in the building):
  A free 60-90 min "Pilot Workshop" offered to HR/L&D contacts.
  Do NOT call it a workshop — call it a "Diagnostic Session" or "Team Stress Audit."
  The lead magnet must demonstrate measurable change in the room (not just awareness).
  Outcome: HR sees the ROI proof before they spend a penny.

STAGE 2 — PAID PILOT (low-risk first engagement):
  A 4-week pilot with one team (8–15 people). Fixed price. No long commitment.
  This is the buying trigger — framed as "risk-free pilot" not "coaching program."
  Price: $3,000–$6,000 for the pilot.

STAGE 3 — ADOPTION PROGRAM (recurring revenue):
  A 6–12 month organisation-wide rollout with train-the-trainer, manager cohorts, or ongoing team sessions.
  This is where the real revenue is: $15,000–$60,000/year per company.

The adoption_funnel field must contain all 3 stages with exact scripts and framing.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
RULE 4 — MARKETING FOR CORPORATE BUYERS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Corporate buyers are NOT on Instagram. They are on:
- LinkedIn (primary): HR directors, L&D managers, Chief People Officers, founders of 50–500 person companies
- HR industry publications and podcasts (SHRM, People Management, HR Brew)
- Conference speaking (HR Tech, Wellbeing at Work, CIPD)
- Warm referrals from current 1:1 clients who work in those companies

LinkedIn content for corporate must:
- Lead with the COST of the problem (burnout, turnover, absenteeism) not the solution
- Use data/statistics, not personal stories
- CTA must offer the free Pilot Workshop / Diagnostic Session — not a "discovery call"

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
RULE 5 — COMPETITORS
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Name 3 REAL people/companies already selling corporate wellness/resilience programs in this niche.
Include their actual pricing if known. These must be real — never invent names.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
GENERAL RULES
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
- Return ONLY valid JSON. No markdown. No code blocks.
- All numeric fields MUST be plain integers — no $ signs, no text.
- Use the coach's exact words from their quiz answers throughout the personal_note and ceiling_narrative.
- Every field must be specific to THEIR niche — never generic wellness platitudes."""


# ─────────────────────────────────────────────────────────────────────────────
# OPENAI CALLS
# ─────────────────────────────────────────────────────────────────────────────

def call_openai(answers: dict) -> dict:
    path = _detect_path(answers)
    if path == "A":
        return _call_openai_path_a(answers)
    if path in ("C", "D"):
        return _call_openai_path_c(answers)
    return _call_openai_path_b(answers)


def _call_openai_path_b(answers: dict) -> dict:
    user_message = f"""
Generate a personalised coaching business plan (Path B — Building) from these quiz answers.

QA1  - Experience type: {answers.get('Q1', answers.get('QA1', ''))}
QA2  - Natural advice area: "{answers.get('Q2', answers.get('QA2', ''))}"
QA3  - Client problem in their exact words: "{answers.get('Q5', answers.get('QA3', ''))}"
QA4  - Ideal client: {answers.get('Q4', answers.get('QA4', ''))}
QA5  - What clients want: "{answers.get('Q6', answers.get('QA5', ''))}"
QA6  - Would not work with: "{answers.get('Q7', answers.get('QA6', ''))}"
QA7  - Current acquisition channels: "{answers.get('Q17', answers.get('QA7', ''))}"
QA8  - Delivery format: {answers.get('Q8', answers.get('QA8', ''))}
QA9  - Duration: {answers.get('Q9', answers.get('QA9', ''))}
QA10 - Pricing comfort: {answers.get('Q10', answers.get('QA10', ''))}
QA11 - Audience size: {answers.get('Q12', answers.get('QA11', ''))}
QA12 - Biggest fear: {answers.get('Q13', answers.get('QA12', ''))}
QA13 - Hours per week: {answers.get('Q14', answers.get('QA13', ''))}
QA14 - Success definition: "{answers.get('Q16', answers.get('QA14', ''))}"
QA15_interest - Open to paid marketing: {answers.get('QA15_interest', 'Yes')}
QA15 - Ad budget: {answers.get('Q15', answers.get('QA15', ''))}

Return this exact JSON structure fully populated:
{{
  "coach_path": "B",
  "offer_sentence": "",
  "offer_layers": [
    {{"layer": "WHO",       "raw": "", "refined": "", "why": ""}},
    {{"layer": "PAIN",      "raw": "", "refined": "", "why": ""}},
    {{"layer": "OUTCOME",   "raw": "", "refined": "", "why": ""}},
    {{"layer": "MECHANISM", "raw": "", "refined": "", "why": ""}},
    {{"layer": "PROOF",     "raw": "", "refined": "", "why": ""}},
    {{"layer": "EXCLUSION", "raw": "", "refined": "", "why": ""}}
  ],
  "funnel": [
    {{"tier": "LEAD MAGNET", "name": "", "format": "", "price": "Free",   "purpose": "", "monthly_clients": 0, "monthly_revenue_low": 0, "monthly_revenue_high": 0}},
    {{"tier": "LOW-TICKET",  "name": "", "format": "", "price": "",       "purpose": "", "monthly_clients": 0, "monthly_revenue_low": 0, "monthly_revenue_high": 0}},
    {{"tier": "MID-TICKET",  "name": "", "format": "", "price": "",       "purpose": "", "monthly_clients": 0, "monthly_revenue_low": 0, "monthly_revenue_high": 0}},
    {{"tier": "FLAGSHIP",    "name": "", "format": "", "price": "",       "purpose": "", "monthly_clients": 0, "monthly_revenue_low": 0, "monthly_revenue_high": 0}}
  ],
  "action_plan": [
    {{"week": "Weeks 1-2",   "focus": "", "actions": "", "milestone": ""}},
    {{"week": "Weeks 3-4",   "focus": "", "actions": "", "milestone": ""}},
    {{"week": "Weeks 5-6",   "focus": "", "actions": "", "milestone": ""}},
    {{"week": "Weeks 7-9",   "focus": "", "actions": "", "milestone": ""}},
    {{"week": "Weeks 10-12", "focus": "", "actions": "", "milestone": ""}}
  ],
  "fear_reframe": {{"fear": "", "truth": "", "action_1": "", "action_2": "", "action_3": ""}},
  "marketing": [
    {{
      "channel": "LinkedIn", "priority": "Primary", "why_this_channel": "",
      "profile_audit": {{}},
      "content_pillars": [
        {{"pillar": "", "purpose": "", "formats": "", "frequency": "", "hooks": ["","",""], "sample_post": ""}}
      ],
      "two_week_calendar": [{{"day": "Mon Wk1", "format": "", "pillar": "", "hook": "", "cta": ""}}],
      "quick_wins": ["","",""],
      "kpis": {{"posting_frequency": "", "engagement_rate_target": "", "connection_growth_per_month": "", "leads_per_month": "", "primary_metric": ""}}
    }}
  ],
  "revenue": [
    {{"program": "", "price_numeric": 0, "year1_low_units": 0, "year1_low_revenue": 0, "year1_high_units": 0, "year1_high_revenue": 0, "year2_units": 0, "year2_revenue": 0}},
    {{"program": "", "price_numeric": 0, "year1_low_units": 0, "year1_low_revenue": 0, "year1_high_units": 0, "year1_high_revenue": 0, "year2_units": 0, "year2_revenue": 0}},
    {{"program": "", "price_numeric": 0, "year1_low_units": 0, "year1_low_revenue": 0, "year1_high_units": 0, "year1_high_revenue": 0, "year2_units": 0, "year2_revenue": 0}},
    {{"program": "", "price_numeric": 0, "year1_low_units": 0, "year1_low_revenue": 0, "year1_high_units": 0, "year1_high_revenue": 0, "year2_units": 0, "year2_revenue": 0}}
  ],
  "revenue_note": "",
  "personal_note": "",
  "trust_channel_benchmarks": {{
    "partnerships": {{"used_now": "", "why_underestimated": "", "estimated_annual_clients": 0, "estimated_annual_revenue": 0, "benchmark_note": "", "first_actions": ["","",""]}},
    "referrals":    {{"used_now": "", "why_underestimated": "", "estimated_annual_clients": 0, "estimated_annual_revenue": 0, "benchmark_note": "", "first_actions": ["","",""]}}
  }},
  "include_paid_funnel": true,
  "competitors": [
    {{"name": "", "url": "", "niche": "", "strategy": "", "content_approach": "", "business_model": "", "flagship_offer": "", "flagship_price": "", "funnel_structure": "", "estimated_revenue": "", "audience_size": "", "strengths": "", "weaknesses": "", "your_edge": ""}},
    {{"name": "", "url": "", "niche": "", "strategy": "", "content_approach": "", "business_model": "", "flagship_offer": "", "flagship_price": "", "funnel_structure": "", "estimated_revenue": "", "audience_size": "", "strengths": "", "weaknesses": "", "your_edge": ""}},
    {{"name": "", "url": "", "niche": "", "strategy": "", "content_approach": "", "business_model": "", "flagship_offer": "", "flagship_price": "", "funnel_structure": "", "estimated_revenue": "", "audience_size": "", "strengths": "", "weaknesses": "", "your_edge": ""}}
  ]
}}
"""
    response = client.chat.completions.create(
        model="gpt-4o", max_tokens=10000,
        response_format={"type": "json_object"},
        messages=[{"role": "system", "content": SYSTEM_PROMPT_B},
                  {"role": "user",   "content": user_message}]
    )
    return json.loads(response.choices[0].message.content)


def _call_openai_path_a(answers: dict) -> dict:
    user_message = f"""
Generate a personalised scale-beyond-1:1 plan (Path A — Established Coach) from these quiz answers.

QB1  - Typical 1:1 client's professional role: "{answers.get('QB1', '')}"
QB2  - Transformation clients describe most often: "{answers.get('QB2', '')}"
QB3  - Has a client ever asked you to work with their team?: {answers.get('QB3', '')}
QB4  - When speaking on stage, who's in the room?: "{answers.get('QB4', '')}"
QB5  - Current 1:1 rate: {answers.get('QB5', '')}
QB6  - Hours per week coaching: {answers.get('QB6', '')}
QB7  - Biggest fear about going corporate or speaking: "{answers.get('QB7', '')}"
QB8  - Preferred new channel: {answers.get('QB8', '')}

Return this exact JSON structure fully populated:
{{
  "coach_path": "A",
  "primary_channel": "corporate",
  "personal_note": "",
  "revenue_ceiling": {{
    "current_hourly_rate": 0,
    "weekly_coaching_hours": 0,
    "monthly_1on1_revenue": 0,
    "annual_1on1_revenue": 0,
    "capacity_ceiling_note": "",
    "corporate_comparison": {{
      "one_engagement_revenue": 0,
      "equivalent_1on1_months": 0,
      "comparison_note": ""
    }}
  }},
  "corporate_offer": {{
    "corporate_pain": "",
    "offer_name": "",
    "format": "",
    "duration": "",
    "buyer_title": "",
    "buying_trigger": "",
    "pricing_logic": "",
    "price_low": 0,
    "price_high": 0,
    "target_companies": [
      {{"name": "", "industry": "", "size": "", "why_now": ""}},
      {{"name": "", "industry": "", "size": "", "why_now": ""}},
      {{"name": "", "industry": "", "size": "", "why_now": ""}}
    ],
    "outreach_one_liner": "",
    "your_edge": ""
  }},
  "speaking_strategy": {{
    "talk_title": "",
    "the_tension": "",
    "audience_pain": "",
    "event_types": [
      {{"type": "", "booking_path": "", "fee_range": "", "notes": ""}},
      {{"type": "", "booking_path": "", "fee_range": "", "notes": ""}},
      {{"type": "", "booking_path": "", "fee_range": "", "notes": ""}}
    ],
    "fee_benchmark_note": "",
    "back_of_room_offer": "",
    "first_outreach_action": ""
  }},
  "action_plan": [
    {{"week": "Weeks 1-4",  "phase": "Translate & Position", "focus": "", "actions": "", "milestone": ""}},
    {{"week": "Weeks 5-8",  "phase": "First Conversations",  "focus": "", "actions": "", "milestone": ""}},
    {{"week": "Weeks 9-12", "phase": "Pilot Engagement",     "focus": "", "actions": "", "milestone": ""}}
  ],
  "fear_reframe": {{
    "fear": "", "truth": "", "action_1": "", "action_2": "", "action_3": ""
  }}
}}
"""
    response = client.chat.completions.create(
        model="gpt-4o", max_tokens=6000,
        response_format={"type": "json_object"},
        messages=[{"role": "system", "content": SYSTEM_PROMPT_A},
                  {"role": "user",   "content": user_message}]
    )
    return json.loads(response.choices[0].message.content)


def _call_openai_path_c(answers: dict) -> dict:
    """
    Path C: Established 1:1 coach with untapped corporate adoption potential.
    Numbers derived from quiz ranges with explicit sourcing labels.
    Corporate offer framed in HR budget language with a 3-stage adoption funnel.
    """
    pricing_range  = answers.get('Q10', answers.get('QA10', '$2,000–$5,000'))
    hours_range    = answers.get('Q14', answers.get('QA13', '15–30 hours'))

    user_message = f"""
Analyse this established 1:1 coach and build their corporate ADOPTION strategy (Path C).

━━━ THEIR QUIZ ANSWERS ━━━
Expertise / niche:           "{answers.get('Q1', answers.get('QA1', ''))}"
Natural advice people seek:  "{answers.get('Q2', answers.get('QA2', ''))}"
Experience level:            {answers.get('Q3', 'Advanced')}
Ideal 1:1 client:            "{answers.get('Q4', answers.get('QA4', ''))}"
Client #1 problem (exact):   "{answers.get('Q5', answers.get('QA5', ''))}"
What clients want most:      "{answers.get('Q6', answers.get('QA5', ''))}"
Would NOT work with:         "{answers.get('Q7', answers.get('QA6', ''))}"
Delivery format:             {answers.get('Q8', answers.get('QA8', ''))}
Program length:              {answers.get('Q9', answers.get('QA9', ''))}
Pricing RANGE (quiz answer): {pricing_range}
Audience size:               {answers.get('Q12', answers.get('QA11', ''))}
Biggest business fear:       "{answers.get('Q13', answers.get('QA12', ''))}"
Hours/week RANGE (quiz):     {hours_range}
Current channels:            "{answers.get('Q17', answers.get('QA7', ''))}"
12-month vision:             "{answers.get('Q16', answers.get('QA14', ''))}"

━━━ INSTRUCTIONS ━━━
1. CEILING: Derive numbers ONLY from the quiz ranges above. Use midpoint, label as estimate.
   - pricing_range_used = "{pricing_range}" → midpoint for calculations
   - hours_range_used   = "{hours_range}" → midpoint for calculations
   - Explain derivation in ceiling_data_source field.

2. CORPORATE PAIN: Connect their niche to ONE specific HR budget line with a real cost figure.
   (burnout, absenteeism, presenteeism, turnover, EAP, L&D, change management)

3. ADOPTION FUNNEL: 3 stages — free Diagnostic Session → paid 4-week Pilot → recurring Adoption Program.
   Each stage needs: name, what happens, price, CTA, success metric.

4. MARKETING: LinkedIn only (B2B). Content must lead with the COST of the problem, not the solution.
   CTA on every post = offer the free Diagnostic Session, not a discovery call.
   Include: 3 content pillars, 3 hooks each, 3 sample posts (150–200 words, publish-ready),
   10-entry 2-week calendar, 3 quick wins (copy-paste ready), KPIs.
   Also include: LinkedIn profile rewrite (headline + about section) in corporate language.

5. OUTREACH: One exact cold email template to an HR Director. Subject line + 4-sentence body.
   Must reference a specific HR pain metric (e.g. "absenteeism costs your sector £X per employee").

6. COMPETITORS: 3 real named companies/coaches selling corporate wellness/resilience in this niche.

Return this exact JSON structure fully populated:
{{
  "coach_path": "C",
  "headline": "",
  "personal_note": "",

  "hidden_ceiling_diagnosis": {{
    "pricing_range_used": "{pricing_range}",
    "hours_range_used": "{hours_range}",
    "estimated_price_per_client": 0,
    "estimated_weekly_coaching_hours": 0,
    "estimated_monthly_max_revenue": 0,
    "estimated_annual_max_revenue": 0,
    "ceiling_data_source": "Estimated from quiz: pricing range midpoint = $X, hours range midpoint = Y hrs/week. Monthly max = X × (Y hrs ÷ program_length_hrs).",
    "ceiling_narrative": "",
    "what_they_dont_know": ""
  }},

  "corporate_pain": {{
    "hr_budget_line": "",
    "cost_of_problem": "",
    "cost_source": "",
    "how_their_niche_maps_to_this_pain": "",
    "what_hr_has_tried_that_didnt_work": "",
    "why_this_coach_is_different": ""
  }},

  "adoption_funnel": {{
    "stage_1": {{
      "name": "Free Diagnostic Session",
      "what_it_is": "",
      "duration": "60–90 minutes",
      "what_happens_in_the_room": "",
      "measurable_outcome_for_hr": "",
      "how_to_position_it": "",
      "cta_to_offer_it": "",
      "price": "Free"
    }},
    "stage_2": {{
      "name": "4-Week Team Pilot",
      "what_it_is": "",
      "duration": "4 weeks",
      "what_happens": "",
      "deliverables": "",
      "success_metric": "",
      "how_to_frame_it_to_hr": "",
      "price_low": 0,
      "price_high": 0,
      "pricing_logic": ""
    }},
    "stage_3": {{
      "name": "Organisation Adoption Program",
      "what_it_is": "",
      "duration": "",
      "format": "",
      "deliverables": "",
      "success_metric": "",
      "renewal_trigger": "",
      "price_low": 0,
      "price_high": 0,
      "pricing_logic": "",
      "annual_recurring_potential": 0
    }}
  }},

  "target_organisations": [
    {{"name": "", "industry": "", "size": "", "hr_pain_signal": "", "why_now": "", "entry_contact_title": ""}},
    {{"name": "", "industry": "", "size": "", "hr_pain_signal": "", "why_now": "", "entry_contact_title": ""}},
    {{"name": "", "industry": "", "size": "", "hr_pain_signal": "", "why_now": "", "entry_contact_title": ""}}
  ],

  "cold_outreach_email": {{
    "subject_line": "",
    "body": "",
    "cta": "",
    "follow_up_day_3": "",
    "follow_up_day_7": ""
  }},

  "revenue_comparison": {{
    "scenario_1on1_only": {{
      "label": "Current 1:1 Only (estimated)",
      "monthly_revenue": 0,
      "annual_revenue": 0,
      "hours_per_week": 0,
      "note": ""
    }},
    "scenario_add_pilot": {{
      "label": "1:1 + One Corporate Pilot/Month",
      "monthly_revenue": 0,
      "annual_revenue": 0,
      "hours_per_week": 0,
      "note": ""
    }},
    "scenario_full_adoption": {{
      "label": "3 Adoption Clients (Year 2)",
      "monthly_revenue": 0,
      "annual_revenue": 0,
      "hours_per_week": 0,
      "note": ""
    }}
  }},

  "linkedin_strategy": {{
    "profile_rewrite": {{
      "headline_before": "",
      "headline_after": "",
      "about_before": "",
      "about_after": "",
      "banner_tip": "",
      "featured_section": ""
    }},
    "content_pillars": [
      {{
        "pillar": "",
        "hr_pain_it_addresses": "",
        "hooks": ["", "", ""],
        "sample_post": "",
        "cta_in_post": "Book the free Diagnostic Session"
      }},
      {{
        "pillar": "",
        "hr_pain_it_addresses": "",
        "hooks": ["", "", ""],
        "sample_post": "",
        "cta_in_post": ""
      }},
      {{
        "pillar": "",
        "hr_pain_it_addresses": "",
        "hooks": ["", "", ""],
        "sample_post": "",
        "cta_in_post": ""
      }}
    ],
    "two_week_calendar": [
      {{"day": "Mon Wk1", "format": "", "pillar": "", "hook": "", "cta": ""}},
      {{"day": "Tue Wk1", "format": "", "pillar": "", "hook": "", "cta": ""}},
      {{"day": "Wed Wk1", "format": "", "pillar": "", "hook": "", "cta": ""}},
      {{"day": "Thu Wk1", "format": "", "pillar": "", "hook": "", "cta": ""}},
      {{"day": "Fri Wk1", "format": "", "pillar": "", "hook": "", "cta": ""}},
      {{"day": "Mon Wk2", "format": "", "pillar": "", "hook": "", "cta": ""}},
      {{"day": "Tue Wk2", "format": "", "pillar": "", "hook": "", "cta": ""}},
      {{"day": "Wed Wk2", "format": "", "pillar": "", "hook": "", "cta": ""}},
      {{"day": "Thu Wk2", "format": "", "pillar": "", "hook": "", "cta": ""}},
      {{"day": "Fri Wk2", "format": "", "pillar": "", "hook": "", "cta": ""}}
    ],
    "quick_wins": ["", "", ""],
    "kpis": {{
      "posting_frequency": "5x/week",
      "connection_target": "",
      "leads_per_month": "",
      "diagnostic_sessions_booked_per_month": "",
      "primary_metric": ""
    }}
  }},

  "action_plan": [
    {{"week": "Weeks 1-2",   "phase": "Translate Your IP",       "focus": "", "actions": "", "milestone": ""}},
    {{"week": "Weeks 3-4",   "phase": "Build the Entry Offer",   "focus": "", "actions": "", "milestone": ""}},
    {{"week": "Weeks 5-6",   "phase": "Warm Outreach Begins",    "focus": "", "actions": "", "milestone": ""}},
    {{"week": "Weeks 7-9",   "phase": "Deliver First Pilot",     "focus": "", "actions": "", "milestone": ""}},
    {{"week": "Weeks 10-12", "phase": "Convert & Systematise",   "focus": "", "actions": "", "milestone": ""}}
  ],

  "fear_reframe": {{
    "fear": "",
    "truth": "",
    "action_1": "",
    "action_2": "",
    "action_3": ""
  }},

  "competitors": [
    {{
      "name": "", "url": "", "niche": "",
      "how_they_sell_to_corporates": "",
      "flagship_corporate_offer": "", "flagship_price": "",
      "audience_size": "", "what_they_do_well": "",
      "gap_you_can_fill": "", "your_edge": ""
    }},
    {{
      "name": "", "url": "", "niche": "",
      "how_they_sell_to_corporates": "",
      "flagship_corporate_offer": "", "flagship_price": "",
      "audience_size": "", "what_they_do_well": "",
      "gap_you_can_fill": "", "your_edge": ""
    }},
    {{
      "name": "", "url": "", "niche": "",
      "how_they_sell_to_corporates": "",
      "flagship_corporate_offer": "", "flagship_price": "",
      "audience_size": "", "what_they_do_well": "",
      "gap_you_can_fill": "", "your_edge": ""
    }}
  ]
}}
"""
    response = client.chat.completions.create(
        model="gpt-4o", max_tokens=12000,
        response_format={"type": "json_object"},
        messages=[{"role": "system", "content": SYSTEM_PROMPT_C},
                  {"role": "user",   "content": user_message}]
    )
    return json.loads(response.choices[0].message.content)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def cl(ws, row, col, val, bold=False, italic=False, color="000000",
       bg=None, size=10, wrap=True, align="left"):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font      = Font(bold=bold, italic=italic, color=color, size=size, name="Arial")
    cell.alignment = Alignment(wrap_text=wrap, vertical="top", horizontal=align)
    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    return cell

def rh(ws, row, h):  ws.row_dimensions[row].height = h
def cw(ws, col, w):  ws.column_dimensions[get_column_letter(col)].width = w

def section_header(ws, row, text, span, bg="1F3864"):
    cl(ws, row, 1, text, bold=True, color="FFFFFF", bg=bg, size=12)
    ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}")
    rh(ws, row, 24)

def col_headers(ws, row, headers, bg="2E75B6"):
    for i, h in enumerate(headers, 1):
        cl(ws, row, i, h, bold=True, color="FFFFFF", bg=bg, size=10, align="center")
    rh(ws, row, 22)

def money_fmt(cell): cell.number_format = '"$"#,##0'
def pct_fmt(cell):   cell.number_format = '0.0%'

LOCK_BG       = "F2F2F2"
LOCK_FG       = "9E9E9E"
LOCK_LABEL_BG = "D9E1F2"
LOCK_LABEL_FG = "1F3864"

# Path C colour palette — warm amber/gold for "opportunity detected"
C_DARK   = "3D2B00"
C_MID    = "7B5200"
C_GOLD   = "F59E0B"
C_LIGHT  = "FEF3C7"
C_ACCENT = "FFFBEB"

def _tease(val: str, keep: int = 45) -> str:
    if not val: return "🔒 [Upgrade to unlock]"
    if len(val) <= keep: return val[:keep] + "  🔒"
    return val[:keep].rstrip() + "…  🔒 [Upgrade to unlock]"

def _lock(ws, row, col, val="", align="left"):
    return cl(ws, row, col, val or "🔒", italic=True, color=LOCK_FG, bg=LOCK_BG, align=align)


# ─────────────────────────────────────────────────────────────────────────────
# PATH C — EXCEL SHEETS
# ─────────────────────────────────────────────────────────────────────────────

def build_c_diagnosis_sheet(ws, data, is_free=False):
    """Sheet 1: Hidden ceiling + corporate pain — the hook."""
    diag = data.get("hidden_ceiling_diagnosis", {})
    corp = data.get("corporate_pain", {})

    section_header(ws, 1, "🔍  YOUR HIDDEN REVENUE CEILING — DIAGNOSED", span=2, bg=C_DARK)
    ws.merge_cells("A2:B2")
    cl(ws, 2, 1, data.get("headline", ""), bold=True, italic=True,
       color=C_DARK, bg=C_LIGHT, size=13, align="center")
    rh(ws, 2, 32)

    # note: keys now use estimated_ prefix from new schema
    diag_rows = [
        ("PRICING RANGE (FROM QUIZ)",    diag.get("pricing_range_used", ""),                   False, False),
        ("HOURS/WEEK RANGE (FROM QUIZ)", diag.get("hours_range_used", ""),                      False, False),
        ("EST. PRICE PER CLIENT",        diag.get("estimated_price_per_client", 0),             False, True),
        ("EST. WEEKLY COACHING HOURS",   diag.get("estimated_weekly_coaching_hours", 0),        False, False),
        ("EST. MONTHLY REVENUE CEILING", diag.get("estimated_monthly_max_revenue", 0),          False, True),
        ("EST. ANNUAL REVENUE CEILING",  diag.get("estimated_annual_max_revenue", 0),           False, True),
        ("HOW THESE WERE CALCULATED",    diag.get("ceiling_data_source", ""),                   False, False),
        ("WHY THIS IS A CEILING",        diag.get("ceiling_narrative", ""),                     False, False),
        ("WHAT YOU DON'T KNOW YET",      diag.get("what_they_dont_know", ""),                   is_free, False),
    ]
    for ri, (label, val, locked, is_money) in enumerate(diag_rows, 4):
        cl(ws, ri, 1, label, bold=True, bg="FDE68A", color=C_DARK, size=11)
        if locked:
            _lock(ws, ri, 2, "🔒 [Unlock the full diagnosis]")
        else:
            c = cl(ws, ri, 2, val, bg=C_ACCENT, size=11)
            if is_money and isinstance(val, int): money_fmt(c)
        rh(ws, ri, 52)

    gap_row = 4 + len(diag_rows) + 1
    section_header(ws, gap_row, "🏢  THE CORPORATE PAIN YOUR NICHE SOLVES", span=2, bg=C_MID)
    corp_rows = [
        ("HR BUDGET LINE",                corp.get("hr_budget_line", ""),               False),
        ("COST OF THE PROBLEM",           corp.get("cost_of_problem", ""),              False),
        ("SOURCE",                        corp.get("cost_source", ""),                  False),
        ("HOW YOUR NICHE MAPS TO IT",     corp.get("how_their_niche_maps_to_this_pain",""), False),
        ("WHAT HR HAS TRIED THAT FAILED", corp.get("what_hr_has_tried_that_didnt_work",""), is_free),
        ("WHY YOU'RE DIFFERENT",          corp.get("why_this_coach_is_different",""),   is_free),
    ]
    for ri, (label, val, locked) in enumerate(corp_rows, gap_row + 2):
        cl(ws, ri, 1, label, bold=True, bg="FDE68A", color=C_DARK, size=10)
        if locked:
            _lock(ws, ri, 2, _tease(val, keep=60))
        else:
            cl(ws, ri, 2, val, bg=C_ACCENT, size=10)
        rh(ws, ri, 55)

    cw(ws, 1, 36); cw(ws, 2, 72)


def build_c_offer_sheet(ws, data, is_free=False):
    """Sheet 2: The 3-stage corporate adoption funnel."""
    funnel = data.get("adoption_funnel", {})
    s1 = funnel.get("stage_1", {})
    s2 = funnel.get("stage_2", {})
    s3 = funnel.get("stage_3", {})
    outreach = data.get("cold_outreach_email", {})

    section_header(ws, 1, "🚀  YOUR 3-STAGE CORPORATE ADOPTION FUNNEL", span=2, bg=C_DARK)

    # Stage 1 — Free Lead Magnet
    row = 3
    section_header(ws, row, f"STAGE 1 (FREE) — {s1.get('name','Free Diagnostic Session')}", span=2, bg=C_MID)
    row += 1
    stage1_rows = [
        ("WHAT IT IS",                s1.get("what_it_is",""),              False),
        ("DURATION",                  s1.get("duration","60–90 minutes"),   False),
        ("WHAT HAPPENS IN THE ROOM",  s1.get("what_happens_in_the_room",""), False),
        ("MEASURABLE OUTCOME FOR HR", s1.get("measurable_outcome_for_hr",""), False),
        ("HOW TO POSITION IT",        s1.get("how_to_position_it",""),      False),
        ("CTA TO OFFER IT",           s1.get("cta_to_offer_it",""),         is_free),
        ("PRICE",                     "Free",                               False),
    ]
    for label, val, locked in stage1_rows:
        cl(ws, row, 1, label, bold=True, bg="FDE68A", color=C_DARK)
        _lock(ws, row, 2, _tease(val, 60)) if locked else cl(ws, row, 2, val, bg=C_ACCENT)
        rh(ws, row, 50); row += 1

    # Stage 2 — Paid Pilot
    row += 1
    section_header(ws, row, f"STAGE 2 (PAID PILOT) — {s2.get('name','4-Week Team Pilot')}", span=2, bg=C_MID)
    row += 1
    stage2_rows = [
        ("WHAT IT IS",            s2.get("what_it_is",""),        False),
        ("DURATION",              s2.get("duration","4 weeks"),   False),
        ("WHAT HAPPENS",          s2.get("what_happens",""),      False),
        ("DELIVERABLES",          s2.get("deliverables",""),      False),
        ("SUCCESS METRIC",        s2.get("success_metric",""),    False),
        ("HOW TO FRAME IT TO HR", s2.get("how_to_frame_it_to_hr",""), is_free),
        ("PRICE RANGE",           f"${s2.get('price_low',0):,} – ${s2.get('price_high',0):,}", is_free),
        ("PRICING LOGIC",         s2.get("pricing_logic",""),    is_free),
    ]
    for label, val, locked in stage2_rows:
        cl(ws, row, 1, label, bold=True, bg="FDE68A", color=C_DARK)
        _lock(ws, row, 2, _tease(val, 60)) if locked else cl(ws, row, 2, val, bg=C_ACCENT)
        rh(ws, row, 50); row += 1

    # Stage 3 — Adoption Program
    row += 1
    section_header(ws, row, f"STAGE 3 (RECURRING) — {s3.get('name','Organisation Adoption Program')}", span=2, bg=C_DARK)
    row += 1
    stage3_rows = [
        ("WHAT IT IS",        s3.get("what_it_is",""),     False),
        ("DURATION",          s3.get("duration",""),        False),
        ("FORMAT",            s3.get("format",""),          False),
        ("DELIVERABLES",      s3.get("deliverables",""),    is_free),
        ("SUCCESS METRIC",    s3.get("success_metric",""),  is_free),
        ("RENEWAL TRIGGER",   s3.get("renewal_trigger",""), is_free),
        ("PRICE RANGE",       f"${s3.get('price_low',0):,} – ${s3.get('price_high',0):,}", is_free),
        ("PRICING LOGIC",     s3.get("pricing_logic",""),   is_free),
        ("ANNUAL RECURRING",  s3.get("annual_recurring_potential", 0), is_free),
    ]
    for label, val, locked in stage3_rows:
        is_money = label == "ANNUAL RECURRING"
        cl(ws, row, 1, label, bold=True, bg="FDE68A", color=C_DARK)
        if locked:
            _lock(ws, row, 2, _tease(str(val), 60))
        else:
            c = cl(ws, row, 2, val, bg=C_ACCENT)
            if is_money and isinstance(val, int): money_fmt(c)
        rh(ws, row, 50); row += 1

    # Cold Outreach Email
    row += 1
    section_header(ws, row, "📧  COLD OUTREACH EMAIL — TO HR DIRECTOR", span=2, bg=C_MID)
    row += 1
    email_rows = [
        ("SUBJECT LINE",    outreach.get("subject_line",""),   False),
        ("BODY",            outreach.get("body",""),           is_free),
        ("CTA",             outreach.get("cta",""),            is_free),
        ("FOLLOW-UP DAY 3", outreach.get("follow_up_day_3",""), is_free),
        ("FOLLOW-UP DAY 7", outreach.get("follow_up_day_7",""), is_free),
    ]
    for label, val, locked in email_rows:
        cl(ws, row, 1, label, bold=True, bg="FDE68A", color=C_DARK)
        _lock(ws, row, 2, _tease(val, 80)) if locked else cl(ws, row, 2, val, bg=C_ACCENT)
        rh(ws, row, 60); row += 1

    # Target organisations
    row += 1
    section_header(ws, row, "🎯  FIRST 3 TARGET ORGANISATIONS", span=3, bg=C_MID)
    row += 1
    col_headers(ws, row, ["ORGANISATION", "INDUSTRY / SIZE", "ENTRY CONTACT + WHY NOW"], bg=C_MID)
    row += 1
    for ti, tgt in enumerate(data.get("target_organisations", [])[:3]):
        bg = C_LIGHT if ti % 2 == 0 else "FFFFFF"
        if is_free:
            cl(ws, row, 1, f"🔒 Target {ti+1}", italic=True, color=LOCK_FG, bg=LOCK_BG)
            _lock(ws, row, 2); _lock(ws, row, 3)
        else:
            cl(ws, row, 1, tgt.get("name",""),    bold=True, bg=bg, color=C_DARK)
            cl(ws, row, 2, f"{tgt.get('industry','')} / {tgt.get('size','')}", bg=bg)
            cl(ws, row, 3, f"{tgt.get('entry_contact_title','')} — {tgt.get('why_now','')} | Signal: {tgt.get('hr_pain_signal','')}", bg=bg)
        rh(ws, row, 55); row += 1

    cw(ws, 1, 30); cw(ws, 2, 72)


def build_c_revenue_comparison_sheet(ws, data, is_free=False):
    """Sheet 3: Side-by-side revenue scenarios."""
    rc = data.get("revenue_comparison", {})
    s1 = rc.get("scenario_1on1_only",    {})
    s2 = rc.get("scenario_add_pilot",    {})
    s3 = rc.get("scenario_full_adoption",{})

    section_header(ws, 1, "📊  REVENUE SCENARIO COMPARISON", span=4, bg=C_DARK)
    col_headers(ws, 3, ["METRIC", "1:1 ONLY (NOW)", "1:1 + ONE PILOT/MONTH", "3 ADOPTION CLIENTS (YR 2)"], bg=C_MID)

    rows_data = [
        ("LABEL",           s1.get("label",""), s2.get("label",""), s3.get("label","")),
        ("MONTHLY REVENUE", s1.get("monthly_revenue",0), s2.get("monthly_revenue",0), s3.get("monthly_revenue",0)),
        ("ANNUAL REVENUE",  s1.get("annual_revenue",0),  s2.get("annual_revenue",0),  s3.get("annual_revenue",0)),
        ("HOURS/WEEK",      s1.get("hours_per_week",0),  s2.get("hours_per_week",0),  s3.get("hours_per_week",0)),
        ("NOTES",           s1.get("note",""),            s2.get("note",""),           s3.get("note","")),
    ]
    for ri, (label, v1, v2, v3) in enumerate(rows_data, 4):
        is_money  = "REVENUE" in label
        locked_s2 = is_free and label not in ("LABEL", "MONTHLY REVENUE")
        locked_s3 = is_free

        cl(ws, ri, 1, label, bold=True, bg="FDE68A", color=C_DARK)
        c2 = cl(ws, ri, 2, v1, bg=C_ACCENT)
        if is_money and isinstance(v1, int): money_fmt(c2)

        if locked_s2:
            _lock(ws, ri, 3, _tease(str(v2), 30))
        else:
            c3 = cl(ws, ri, 3, v2, bg="D1FAE5", bold=(label == "ANNUAL REVENUE"))
            if is_money and isinstance(v2, int): money_fmt(c3)

        if locked_s3:
            _lock(ws, ri, 4, "🔒 [Full Plan Only]")
        else:
            c4 = cl(ws, ri, 4, v3, bg="FEF9C3", bold=(label == "ANNUAL REVENUE"))
            if is_money and isinstance(v3, int): money_fmt(c4)

        rh(ws, ri, 50)

    cw(ws, 1, 24); cw(ws, 2, 30); cw(ws, 3, 34); cw(ws, 4, 34)


def build_c_marketing_sheet(ws, data, is_free=False):
    """Sheet 4: LinkedIn-only B2B marketing strategy for the corporate offer."""
    li = data.get("linkedin_strategy", {})
    audit = li.get("profile_rewrite", {})

    section_header(ws, 1, "📣  LINKEDIN STRATEGY FOR CORPORATE BUYERS", span=5, bg=C_DARK)
    ws.merge_cells("A2:E2")
    cl(ws, 2, 1,
       "Corporate buyers are NOT on Instagram. This is your LinkedIn strategy to reach HR Directors, L&D Managers, and Chief People Officers.",
       italic=True, bg=C_LIGHT, color=C_DARK, size=10)
    rh(ws, 2, 28)

    # Profile rewrite
    row = 4
    section_header(ws, row, "👤  PROFILE REWRITE — CORPORATE LANGUAGE", span=5, bg=C_MID)
    row += 1
    profile_rows = [
        ("HEADLINE (BEFORE)", audit.get("headline_before",""), False),
        ("HEADLINE (AFTER)",  audit.get("headline_after",""),  False),
        ("ABOUT (BEFORE)",    audit.get("about_before",""),    False),
        ("ABOUT (AFTER)",     audit.get("about_after",""),     is_free),
        ("BANNER TIP",        audit.get("banner_tip",""),      False),
        ("FEATURED SECTION",  audit.get("featured_section",""), is_free),
    ]
    for label, val, locked in profile_rows:
        cl(ws, row, 1, label, bold=True, bg="FDE68A", color=C_DARK, size=9)
        ws.merge_cells(f"B{row}:E{row}")
        if locked:
            cl(ws, row, 2, _tease(val, 80), italic=True, color=LOCK_FG, bg=LOCK_BG, size=9)
        else:
            cl(ws, row, 2, val, bg=C_ACCENT, size=9)
        rh(ws, row, 55); row += 1

    # Content pillars
    row += 1
    section_header(ws, row, "📌  CONTENT PILLARS — HR PAIN FIRST, SOLUTION SECOND", span=5, bg=C_MID)
    row += 1
    for pi, pillar in enumerate(li.get("content_pillars", [])):
        bg_p = C_LIGHT if pi % 2 == 0 else "FFFFFF"
        cl(ws, row, 1, pillar.get("pillar",""),              bold=True, bg="FDE68A", color=C_DARK, size=9)
        ws.merge_cells(f"B{row}:C{row}")
        cl(ws, row, 2, f"HR pain: {pillar.get('hr_pain_it_addresses','')}",  bg=bg_p, size=9)
        ws.merge_cells(f"D{row}:E{row}")
        hooks_text = "\n".join(f"▸ {h}" for h in pillar.get("hooks", []))
        cl(ws, row, 4, hooks_text, bg=bg_p, size=9)
        rh(ws, row, 60); row += 1

        post = pillar.get("sample_post","")
        cta  = pillar.get("cta_in_post","")
        cl(ws, row, 1, "SAMPLE POST", bold=True, bg="FDE68A", color=C_DARK, size=9)
        ws.merge_cells(f"B{row}:E{row}")
        full_post = f"{post}\n\n📌 CTA: {cta}" if cta else post
        if is_free:
            cl(ws, row, 2, (full_post[:90]+"…  🔒") if len(full_post)>90 else full_post+"  🔒",
               italic=True, color=LOCK_FG, bg=LOCK_BG, size=9)
            rh(ws, row, 40)
        else:
            cl(ws, row, 2, full_post, bg=bg_p, size=9)
            rh(ws, row, max(80, min(220, len(full_post)//2)))
        row += 1

    # 2-week calendar
    row += 1
    section_header(ws, row, "📅  2-WEEK POSTING CALENDAR", span=5, bg=C_MID)
    row += 1
    col_headers(ws, row, ["DAY", "FORMAT", "PILLAR", "HOOK", "CTA"], bg=C_MID)
    row += 1
    for entry in li.get("two_week_calendar", []):
        bg = C_LIGHT if row % 2 == 0 else "FFFFFF"
        cl(ws, row, 1, entry.get("day",""),    bold=True, bg="FDE68A", color=C_DARK, size=9)
        cl(ws, row, 2, entry.get("format",""), bg=bg, size=9)
        cl(ws, row, 3, entry.get("pillar",""), bg=bg, size=9)
        cl(ws, row, 4, entry.get("hook",""),   bg=bg, size=9)
        cl(ws, row, 5, entry.get("cta",""),    bg=bg, size=9)
        rh(ws, row, 40); row += 1

    # Quick wins
    row += 1
    section_header(ws, row, "⚡  QUICK WINS — DO IN 48 HRS", span=5, bg=C_MID)
    row += 1
    for qi, win in enumerate(li.get("quick_wins",[]), 1):
        cl(ws, row, 1, f"#{qi}", bold=True, bg="FDE68A", color=C_DARK, size=9, align="center")
        ws.merge_cells(f"B{row}:E{row}")
        if is_free and qi > 1:
            cl(ws, row, 2, _tease(win, 60), italic=True, color=LOCK_FG, bg=LOCK_BG, size=9)
        else:
            cl(ws, row, 2, win, bg=C_ACCENT, size=9)
        rh(ws, row, 45); row += 1

    # KPIs
    row += 1
    kpis = li.get("kpis", {})
    section_header(ws, row, "📈  SUCCESS METRICS", span=5, bg=C_MID)
    row += 1
    for kk, kv in kpis.items():
        cl(ws, row, 1, kk.replace("_"," ").title(), bold=True, bg="FDE68A", color=C_DARK, size=9)
        ws.merge_cells(f"B{row}:E{row}")
        cl(ws, row, 2, str(kv), bg=C_ACCENT, size=9)
        rh(ws, row, 28); row += 1

    cw(ws, 1, 22); cw(ws, 2, 24); cw(ws, 3, 22); cw(ws, 4, 30); cw(ws, 5, 24)


def build_c_action_sheet(ws, data, is_free=False):
    """Sheet 5: 90-day action plan."""
    section_header(ws, 1, "📅  YOUR 90-DAY EXPANSION ACTION PLAN", span=5, bg=C_DARK)
    col_headers(ws, 3, ["WEEK", "PHASE", "FOCUS", "ACTIONS", "MILESTONE"], bg=C_MID)
    phase_colors = [C_DARK, C_MID, "2E5B8A", "1F6B45", "7B2D00"]
    action_plan = data.get("action_plan", [])

    for ri, item in enumerate(action_plan, 4):
        idx    = ri - 4
        locked = is_free and idx >= 2
        bg_ph  = phase_colors[idx] if idx < len(phase_colors) else C_MID

        cl(ws, ri, 1, item.get("week",""),  bold=True, bg=bg_ph, color="FFFFFF")
        cl(ws, ri, 2, item.get("phase",""), bold=True, bg="FDE68A", color=C_DARK)

        if locked:
            _lock(ws, ri, 3, "🔒 [Unlock Expansion Strategy]")
            _lock(ws, ri, 4, _tease(item.get("actions",""), keep=55))
            _lock(ws, ri, 5, _tease(item.get("milestone",""), keep=30))
        else:
            cl(ws, ri, 3, item.get("focus",""),     bold=True, bg=C_ACCENT)
            cl(ws, ri, 4, item.get("actions",""),   bg="FFFFFF")
            cl(ws, ri, 5, item.get("milestone",""), italic=True, color=C_DARK, bg=C_LIGHT)

        rh(ws, ri, 70)

    legend_row = 4 + len(action_plan) + 1
    ws.merge_cells(f"A{legend_row}:E{legend_row}")
    cl(ws, legend_row, 1, "Weeks 1–4 fully visible.  🔒 Weeks 5–12 unlocked in the Full Expansion Plan.",
       italic=True, bg=C_LIGHT, color=C_DARK, size=9)
    rh(ws, legend_row, 18)
    cw(ws, 1, 14); cw(ws, 2, 22); cw(ws, 3, 28); cw(ws, 4, 44); cw(ws, 5, 30)


def build_c_competitor_sheet(ws, data, is_free=False):
    """Sheet 6: Competitors already monetising this same niche."""
    section_header(ws, 1, "🕵  WHO IS ALREADY MONETISING YOUR NICHE AT SCALE", span=3, bg=C_DARK)
    ws.merge_cells("A2:C2")
    cl(ws, 2, 1, "These are NOT your 1:1 competitors. These are coaches/consultants already selling the EXPANSION OFFER in your niche.",
       italic=True, bg=C_LIGHT, color=C_DARK, size=9)
    rh(ws, 2, 28)

    COMP_COLORS = [("B45309","FEF3C7"), ("1D4ED8","DBEAFE"), ("065F46","D1FAE5")]
    FIELD_LABELS = [
        ("HOW THEY SELL TO CORPORATES",  "how_they_sell_to_corporates"),
        ("FLAGSHIP CORPORATE OFFER",     "flagship_corporate_offer"),
        ("FLAGSHIP PRICE",               "flagship_price"),
        ("AUDIENCE SIZE",                "audience_size"),
        ("WHAT THEY DO WELL",            "what_they_do_well"),
        ("GAP YOU CAN FILL",             "gap_you_can_fill"),
        ("YOUR EDGE",                    "your_edge"),
    ]
    current_row = 4
    for idx, comp in enumerate(data.get("competitors", [])[:3]):
        hbg, bbg = COMP_COLORS[idx]
        is_locked_comp = is_free and idx > 0
        label = (f"🔒 COMPETITOR {idx+1} — [Unlock in Full Plan]"
                 if is_locked_comp else
                 f"COMPETITOR {idx+1} — {comp.get('name','')}")
        cl(ws, current_row, 1, label, bold=True, color="FFFFFF",
           bg=LOCK_FG if is_locked_comp else hbg, size=11)
        ws.merge_cells(f"A{current_row}:C{current_row}")
        rh(ws, current_row, 22); current_row += 1

        col_headers(ws, current_row, ["FIELD", "DETAIL", "YOUR NOTES"],
                    bg=LOCK_FG if is_locked_comp else hbg)
        current_row += 1

        for label_text, key in FIELD_LABELS:
            value = comp.get(key, "")
            locked_field = is_locked_comp or (is_free and key not in ("flagship_offer", "what_they_do_well", "your_edge"))
            alt_bg = bbg if current_row % 2 == 0 else "FFFFFF"

            cl(ws, current_row, 1, label_text, bold=True, bg="FDE68A", color=C_DARK)
            if key == "your_edge" and not is_locked_comp:
                cl(ws, current_row, 2, value, bold=True, bg=C_DARK, color="FFD700", size=10)
                cl(ws, current_row, 3, "", bg=C_DARK)
            elif locked_field:
                cl(ws, current_row, 2, _tease(value, keep=40), italic=True, color=LOCK_FG, bg=LOCK_BG)
                _lock(ws, current_row, 3, "🔒")
            else:
                cl(ws, current_row, 2, value, bg=alt_bg)
                cl(ws, current_row, 3, "", bg=alt_bg)
            rh(ws, current_row, 50); current_row += 1
        current_row += 1

    cw(ws, 1, 30); cw(ws, 2, 52); cw(ws, 3, 40)


def build_c_fear_sheet(ws, data, is_free=False):
    section_header(ws, 1, "😨  YOUR FEAR ABOUT EXPANDING — REFRAMED", span=2, bg=C_DARK)
    fear = data.get("fear_reframe", {})
    rows = [
        ("THE FEAR",           fear.get("fear",""),     "C00000", "FFE0E0", False),
        ("THE REAL TRUTH",     fear.get("truth",""),    C_DARK,   C_LIGHT,  is_free),
        ("ACTION THIS WEEK 1", fear.get("action_1",""), C_MID,    C_ACCENT, is_free),
        ("ACTION THIS WEEK 2", fear.get("action_2",""), C_MID,    C_ACCENT, is_free),
        ("ACTION THIS WEEK 3", fear.get("action_3",""), C_MID,    C_ACCENT, is_free),
    ]
    for ri, (label, val, label_color, val_bg, locked) in enumerate(rows, 3):
        cl(ws, ri, 1, label, bold=True, bg="FDE68A", color=label_color, size=11)
        if locked:
            cl(ws, ri, 2, _tease(val, keep=55), italic=True, color=LOCK_FG, bg=LOCK_BG, size=11)
        else:
            cl(ws, ri, 2, val, bg=val_bg, size=11)
        rh(ws, ri, 60)
    cw(ws, 1, 22); cw(ws, 2, 70)


def build_c_upgrade_sheet(ws, payment_url: str = ""):
    section_header(ws, 1, "🔒  UNLOCK YOUR FULL EXPANSION PLAN", span=2, bg=C_DARK)
    locked_sections = [
        ("🔍 Hidden Ceiling — Full Diagnosis",
         "The exact numbers showing what your 1:1 practice can never earn, no matter how full your calendar gets."),
        ("💡 Untapped Potential — Market Evidence",
         "Real data on who is already buying this in your niche and what they're paying."),
        ("🚀 Expansion Offer — Buyer Title & Buying Trigger",
         "Who has budget authority and the exact moment they're ready to spend it."),
        ("🚀 Expansion Offer — Pricing Logic & Outreach One-Liner",
         "How to anchor your price and the exact first sentence to get a reply."),
        ("🚀 Expansion Offer — First 3 Targets",
         "Three real, named organisations or event producers to approach this month."),
        ("📊 Revenue Comparison — Year 2 Scenario",
         "What fully scaled looks like: revenue, hours, and the inflection point where it clicks."),
        ("📣 Marketing for New Offer — Full Strategy",
         "2-week calendar, sample posts, and KPIs for every channel that reaches expansion buyers."),
        ("📅 90-Day Plan — Weeks 5–12",
         "Warm outreach sequence, pilot delivery, and first close script."),
        ("🕵 Competitor Analysis — All 3",
         "Full breakdown of every competitor: pricing, funnel, their gaps, your edge."),
        ("😨 Fear Reframe (Full)",
         "The real truth behind your hesitation + 3 specific actions to take this week."),
    ]
    for ri, (label, desc) in enumerate(locked_sections, 3):
        cl(ws, ri, 1, f"🔒  {label}", bold=True, bg="FDE68A", color=C_DARK, size=11)
        cl(ws, ri, 2, f"{desc}\n\n➡  Unlock in the Full Expansion Plan.", bg="FFF8E7", color="856404")
        rh(ws, ri, 70)

    cta_row = 3 + len(locked_sections) + 1
    ws.merge_cells(f"A{cta_row}:B{cta_row}")
    cl(ws, cta_row, 1,
       "✨  UNLOCK YOUR FULL EXPANSION PLAN  ✨\n\n"
       "You've seen your hidden ceiling and the shape of your expansion offer.\n"
       "The full plan gives you the exact buyer, the trigger, the pricing logic,\n"
       "the outreach email, competitor intelligence, and your complete 90-day roadmap.",
       bold=True, bg=C_DARK, color="FFFFFF", size=12)
    rh(ws, cta_row, 120)

    link_row = cta_row + 1
    ws.merge_cells(f"A{link_row}:B{link_row}")
    tag_cell = cl(ws, link_row, 1, "👉  CLICK HERE TO UNLOCK YOUR FULL EXPANSION PLAN  →",
                  bold=True, bg=C_GOLD, color=C_DARK, size=14, align="center")
    if payment_url:
        tag_cell.hyperlink = payment_url
    rh(ws, link_row, 40)
    cw(ws, 1, 32); cw(ws, 2, 72)


# ─────────────────────────────────────────────────────────────────────────────
# PATH A — EXCEL SHEETS
# ─────────────────────────────────────────────────────────────────────────────

def build_a_revenue_ceiling_sheet(ws, data, is_free=False):
    rc = data.get("revenue_ceiling", {})
    cc = rc.get("corporate_comparison", {})
    section_header(ws, 1, "REVENUE CEILING DIAGNOSIS", span=2, bg="1a3a2a")

    rows = [
        ("CURRENT HOURLY RATE",          rc.get("current_hourly_rate", 0),    False, True),
        ("WEEKLY COACHING HOURS",         rc.get("weekly_coaching_hours", 0),  False, False),
        ("MONTHLY 1:1 REVENUE",           rc.get("monthly_1on1_revenue", 0),   False, True),
        ("ANNUAL 1:1 REVENUE",            rc.get("annual_1on1_revenue", 0),    False, True),
        ("THE CEILING",                   rc.get("capacity_ceiling_note", ""), False, False),
        ("ONE CORPORATE ENGAGEMENT",      cc.get("one_engagement_revenue", 0), is_free, True),
        ("EQUIVALENT 1:1 MONTHS",         cc.get("equivalent_1on1_months", 0), is_free, False),
        ("WHY THIS MATTERS",              cc.get("comparison_note", ""),       is_free, False),
    ]
    for ri, (label, val, locked, is_money) in enumerate(rows, 3):
        cl(ws, ri, 1, label, bold=True, bg="D9E1F2", color="1a3a2a", size=11)
        if locked:
            _lock(ws, ri, 2, "🔒 [Upgrade to see corporate comparison]")
        else:
            c = cl(ws, ri, 2, val, bg="F0FFF4", size=11)
            if is_money and isinstance(val, int): money_fmt(c)
        rh(ws, ri, 48)
    cw(ws, 1, 30); cw(ws, 2, 65)


def build_a_corporate_offer_sheet(ws, data, is_free=False):
    co = data.get("corporate_offer", {})
    section_header(ws, 1, "CORPORATE ADOPTION OFFER", span=2, bg="1a3a2a")

    meta_rows = [
        ("CORPORATE PAIN",       co.get("corporate_pain", ""),       False),
        ("OFFER NAME",           co.get("offer_name", ""),           False),
        ("FORMAT",               co.get("format", ""),               False),
        ("DURATION",             co.get("duration", ""),             False),
        ("BUYER TITLE",          co.get("buyer_title", ""),          is_free),
        ("BUYING TRIGGER",       co.get("buying_trigger", ""),       is_free),
        ("PRICING LOGIC",        co.get("pricing_logic", ""),        is_free),
        ("PRICE RANGE",          f"${co.get('price_low',0):,} – ${co.get('price_high',0):,}", is_free),
        ("YOUR EDGE",            co.get("your_edge", ""),            False),
        ("OUTREACH ONE-LINER",   co.get("outreach_one_liner", ""),   is_free),
    ]
    row = 3
    for label, val, locked in meta_rows:
        cl(ws, row, 1, label, bold=True, bg="D9E1F2", color="1a3a2a")
        if locked:
            _lock(ws, row, 2, _tease(val, keep=50))
        else:
            cl(ws, row, 2, val, bg="F0FFF4")
        rh(ws, row, 50); row += 1

    row += 1
    section_header(ws, row, "TARGET COMPANIES", span=4, bg="2d6648")
    row += 1
    col_headers(ws, row, ["COMPANY", "INDUSTRY", "SIZE", "WHY NOW"], bg="2d6648")
    row += 1
    for ci, comp in enumerate(co.get("target_companies", [])[:3]):
        bg = "EAF3ED" if ci % 2 == 0 else "FFFFFF"
        if is_free:
            cl(ws, row, 1, f"🔒 Company {ci+1}", italic=True, color=LOCK_FG, bg=LOCK_BG)
            for c in range(2, 5): _lock(ws, row, c)
        else:
            cl(ws, row, 1, comp.get("name", ""),     bold=True, bg=bg, color="1a3a2a")
            cl(ws, row, 2, comp.get("industry", ""), bg=bg)
            cl(ws, row, 3, comp.get("size", ""),     bg=bg, align="center")
            cl(ws, row, 4, comp.get("why_now", ""),  bg=bg)
        rh(ws, row, 45); row += 1

    cw(ws, 1, 26); cw(ws, 2, 30); cw(ws, 3, 18); cw(ws, 4, 40)


def build_a_speaking_sheet(ws, data, is_free=False):
    ss = data.get("speaking_strategy", {})
    section_header(ws, 1, "SPEAKING STRATEGY", span=2, bg="1a3a2a")

    meta_rows = [
        ("TALK TITLE",             ss.get("talk_title", ""),           False),
        ("THE TENSION",            ss.get("the_tension", ""),          False),
        ("AUDIENCE PAIN",          ss.get("audience_pain", ""),        False),
        ("FEE BENCHMARK",          ss.get("fee_benchmark_note", ""),   is_free),
        ("BACK-OF-ROOM OFFER",     ss.get("back_of_room_offer", ""),   is_free),
        ("FIRST OUTREACH ACTION",  ss.get("first_outreach_action", ""), is_free),
    ]
    row = 3
    for label, val, locked in meta_rows:
        cl(ws, row, 1, label, bold=True, bg="D9E1F2", color="1a3a2a")
        if locked:
            _lock(ws, row, 2, _tease(val, keep=55))
        else:
            cl(ws, row, 2, val, bg="F0FFF4")
        rh(ws, row, 50); row += 1

    row += 1
    section_header(ws, row, "3 EVENT TYPES TO TARGET", span=4, bg="2d6648")
    row += 1
    col_headers(ws, row, ["EVENT TYPE", "BOOKING PATH", "FEE RANGE", "NOTES"], bg="2d6648")
    row += 1
    for ei, evt in enumerate(ss.get("event_types", [])[:3]):
        bg = "EAF3ED" if ei % 2 == 0 else "FFFFFF"
        if is_free:
            cl(ws, row, 1, evt.get("type", ""), bold=True, bg=bg, color="1a3a2a")
            for c in range(2, 5): _lock(ws, row, c)
        else:
            cl(ws, row, 1, evt.get("type", ""),         bold=True, bg=bg, color="1a3a2a")
            cl(ws, row, 2, evt.get("booking_path", ""), bg=bg)
            cl(ws, row, 3, evt.get("fee_range", ""),    bg=bg, align="center")
            cl(ws, row, 4, evt.get("notes", ""),        bg=bg)
        rh(ws, row, 45); row += 1

    cw(ws, 1, 26); cw(ws, 2, 36); cw(ws, 3, 18); cw(ws, 4, 40)


def build_a_action_sheet(ws, data, is_free=False):
    section_header(ws, 1, "90-DAY EXPANSION PLAN", span=5, bg="1a3a2a")
    col_headers(ws, 3, ["WEEK", "PHASE", "FOCUS", "ACTIONS", "MILESTONE"], bg="2d6648")
    phase_colors = ["2d6648", "3a8a5a", "4db870"]
    action_plan = data.get("action_plan", [])

    for ri, item in enumerate(action_plan, 4):
        idx    = ri - 4
        locked = is_free and idx >= 1
        bg_ph  = phase_colors[idx] if idx < len(phase_colors) else "2d6648"

        cl(ws, ri, 1, item.get("week", ""),    bold=True, bg=bg_ph, color="FFFFFF")
        cl(ws, ri, 2, item.get("phase", ""),   bold=True, bg="EAF3ED", color="1a3a2a")

        if locked:
            _lock(ws, ri, 3, "🔒 [Unlock Strategy]")
            _lock(ws, ri, 4, _tease(item.get("actions", ""), keep=55))
            _lock(ws, ri, 5, _tease(item.get("milestone", ""), keep=30))
        else:
            cl(ws, ri, 3, item.get("focus", ""),     bold=True, bg="F0FFF4")
            cl(ws, ri, 4, item.get("actions", ""))
            cl(ws, ri, 5, item.get("milestone", ""), italic=True, color="1a3a2a", bg="EAF3ED")

        rh(ws, ri, 70)

    legend_row = 4 + len(action_plan) + 1
    ws.merge_cells(f"A{legend_row}:E{legend_row}")
    cl(ws, legend_row, 1,
       "Weeks 1–4 are fully unlocked.  🔒 Weeks 5–12 are in the Full Plan.",
       italic=True, bg="EAF3ED", color="1a3a2a", size=9)
    rh(ws, legend_row, 18)
    cw(ws, 1, 14); cw(ws, 2, 22); cw(ws, 3, 26); cw(ws, 4, 44); cw(ws, 5, 30)


def build_a_fear_sheet(ws, data, is_free=False):
    section_header(ws, 1, "YOUR FEAR REFRAME", span=2, bg="1a3a2a")
    fear = data.get("fear_reframe", {})
    rows = [
        ("THE FEAR",           fear.get("fear", ""),     "C00000", "FFE0E0", False),
        ("THE REAL TRUTH",     fear.get("truth", ""),    "1a3a2a", "EAF3ED", is_free),
        ("ACTION THIS WEEK 1", fear.get("action_1", ""), "2d6648", "F0FFF4", is_free),
        ("ACTION THIS WEEK 2", fear.get("action_2", ""), "2d6648", "F0FFF4", is_free),
        ("ACTION THIS WEEK 3", fear.get("action_3", ""), "2d6648", "F0FFF4", is_free),
    ]
    for ri, (label, val, label_color, val_bg, locked) in enumerate(rows, 3):
        cl(ws, ri, 1, label, bold=True, bg="D9E1F2", color=label_color, size=11)
        if locked:
            cl(ws, ri, 2, _tease(val, keep=55), italic=True, color=LOCK_FG, bg=LOCK_BG, size=11)
        else:
            cl(ws, ri, 2, val, bg=val_bg, size=11)
        rh(ws, ri, 60)
    cw(ws, 1, 22); cw(ws, 2, 70)


def build_a_upgrade_sheet(ws, payment_url: str = ""):
    section_header(ws, 1, "🔒 UNLOCK YOUR FULL EXPANSION PLAN", span=2, bg="1a3a2a")
    locked_sections = [
        ("📊 Revenue Ceiling — Corporate Comparison",
         "How much one corporate engagement is worth vs. one month of 1:1 coaching."),
        ("🏢 Corporate Offer — Buyer Title & Buying Trigger",
         "The exact job title with budget authority and the signal that tells you they're ready NOW."),
        ("🏢 Corporate Offer — Pricing Logic & 3 Target Companies",
         "How to anchor your price to cost-avoided, plus three real companies to approach this month."),
        ("🏢 Corporate Offer — Outreach One-Liner",
         "The exact first sentence of your cold email — written in their language, not yours."),
        ("🎤 Speaking — Fee Benchmarks & Event Targets",
         "What corporate all-hands actually pay vs. conference fees."),
        ("🎤 Speaking — Back-of-Room Offer & First Outreach Action",
         "How the talk becomes the top of your corporate funnel."),
        ("📅 90-Day Plan — Weeks 5–12",
         "The warm outreach sequence, pilot offer structure, and first engagement close."),
        ("😨 Fear Reframe (Full)",
         "The real truth behind your fear about going corporate + 3 specific actions."),
    ]
    for ri, (label, desc) in enumerate(locked_sections, 3):
        cl(ws, ri, 1, f"🔒  {label}", bold=True, bg="D9E1F2", color="1a3a2a", size=11)
        cl(ws, ri, 2, f"{desc}\n\n➡  Unlock in the Full Expansion Plan.", bg="FFF8E7", color="856404")
        rh(ws, ri, 70)

    cta_row = 3 + len(locked_sections) + 1
    ws.merge_cells(f"A{cta_row}:B{cta_row}")
    cl(ws, cta_row, 1,
       "✨  UNLOCK YOUR FULL EXPANSION PLAN  ✨\n\n"
       "You can see your Revenue Ceiling and the shape of your Corporate Offer.\n"
       "The full plan gives you the exact buyer, buying trigger, pricing logic,\n"
       "3 target companies, the outreach email, and your complete 90-day roadmap.",
       bold=True, bg="1a3a2a", color="FFFFFF", size=12)
    rh(ws, cta_row, 120)

    link_row = cta_row + 1
    ws.merge_cells(f"A{link_row}:B{link_row}")
    tag_cell = cl(ws, link_row, 1, "👉  CLICK HERE TO UNLOCK YOUR FULL PLAN  →",
                  bold=True, bg="FFD700", color="1a3a2a", size=14, align="center")
    if payment_url:
        tag_cell.hyperlink = payment_url
    rh(ws, link_row, 40)
    cw(ws, 1, 32); cw(ws, 2, 72)


# ─────────────────────────────────────────────────────────────────────────────
# PATH B — EXCEL SHEETS
# ─────────────────────────────────────────────────────────────────────────────

def build_b_offer_sheet(ws, data, is_premium=False):
    if is_premium:
        ws.merge_cells("A1:D1")
        cl(ws, 1, 1, "🏆 WELCOME TO YOUR FULL COACHING PLAN — LET'S BUILD!", bold=True, bg="1c1c4a", color="FFD700", size=14, align="center")
        rh(ws, 1, 35)
        offset = 1
    else:
        offset = 0
    section_header(ws, 1 + offset, "YOUR PERSONALISED COACHING OFFER", span=4)
    ws.merge_cells(f"A{3+offset}:D{3+offset}")
    cl(ws, 3+offset, 1, "YOUR ONE-SENTENCE OFFER", bold=True, bg="2E75B6", color="FFFFFF", size=11)
    ws.merge_cells(f"A{4+offset}:D{4+offset}")
    cl(ws, 4+offset, 1, data.get("offer_sentence", ""), bold=True, color="1F3864", size=12, bg="EBF3FB")
    rh(ws, 4+offset, 55)
    if data.get("personal_note"):
        ws.merge_cells(f"A{6+offset}:D{6+offset}")
        cl(ws, 6+offset, 1, f"Note: {data['personal_note']}", italic=True, bg="E2EFDA", color="375623")
        rh(ws, 6+offset, 40)
    section_header(ws, 8+offset, "OFFER LAYERS", span=4, bg="2E75B6")
    col_headers(ws, 9+offset, ["LAYER", "RAW INPUT", "REFINED", "WHY THIS SELLS"])
    for ri, layer in enumerate(data.get("offer_layers", []), 10+offset):
        bg = "F7F9FC" if ri % 2 == 0 else "FFFFFF"
        cl(ws, ri, 1, layer.get("layer", ""), bold=True, bg="D9E1F2", color="1F3864")
        cl(ws, ri, 2, layer.get("raw", ""), bg=bg)
        cl(ws, ri, 3, layer.get("refined", ""), bg=bg)
        cl(ws, ri, 4, layer.get("why", ""), bg=bg)
        rh(ws, ri, 65)
    cw(ws, 1, 18); cw(ws, 2, 35); cw(ws, 3, 42); cw(ws, 4, 40)


def build_b_funnel_sheet(ws, data):
    section_header(ws, 1, "YOUR RECOMMENDED FUNNEL", span=8)
    col_headers(ws, 3, ["TIER", "NAME", "FORMAT", "PRICE", "PURPOSE", "EST. MONTHLY CLIENTS", "MONTHLY REV (LOW)", "MONTHLY REV (HIGH)"])
    colors = {"LEAD MAGNET": "70AD47", "LOW-TICKET": "4472C4", "MID-TICKET": "ED7D31", "FLAGSHIP": "C00000"}
    for ri, item in enumerate(data.get("funnel", []), 4):
        tier = item.get("tier", ""); bg = colors.get(tier, "FFFFFF")
        cl(ws, ri, 1, tier, bold=True, bg=bg, color="FFFFFF")
        cl(ws, ri, 2, item.get("name", ""))
        cl(ws, ri, 3, item.get("format", ""))
        cl(ws, ri, 4, item.get("price", ""), bold=True, color="1F3864")
        cl(ws, ri, 5, item.get("purpose", ""))
        c6 = cl(ws, ri, 6, item.get("monthly_clients", 0), align="center")
        c7 = cl(ws, ri, 7, item.get("monthly_revenue_low", 0),  align="right")
        c8 = cl(ws, ri, 8, item.get("monthly_revenue_high", 0), align="right")
        money_fmt(c7); money_fmt(c8)
        rh(ws, ri, 50)
    tr = 4 + len(data.get("funnel", []))
    cl(ws, tr, 1, "MONTHLY TOTALS", bold=True, bg="1F3864", color="FFFFFF")
    for col in [2, 3, 4]: cl(ws, tr, col, "", bg="D9E1F2")
    cl(ws, tr, 5, "", bg="1F3864")
    for col_idx, col_letter in [(7, "G"), (8, "H")]:
        c = ws.cell(row=tr, column=col_idx, value=f"=SUM({col_letter}4:{col_letter}{tr-1})")
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        c.alignment = Alignment(horizontal="right", vertical="top")
        money_fmt(c)
    c6t = ws.cell(row=tr, column=6, value=f"=SUM(F4:F{tr-1})")
    c6t.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    c6t.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    c6t.alignment = Alignment(horizontal="center", vertical="top")
    rh(ws, tr, 22)
    cw(ws, 1, 16); cw(ws, 2, 26); cw(ws, 3, 22); cw(ws, 4, 14); cw(ws, 5, 38); cw(ws, 6, 20); cw(ws, 7, 18); cw(ws, 8, 18)


def build_b_action_sheet(ws, data):
    section_header(ws, 1, "YOUR 90-DAY ACTION PLAN", span=5)
    col_headers(ws, 3, ["WEEK", "FOCUS", "ACTIONS", "MILESTONE", "STATUS"])
    phase_colors = ["4472C4", "4472C4", "ED7D31", "ED7D31", "C00000"]
    for ri, item in enumerate(data.get("action_plan", []), 4):
        bg = phase_colors[ri-4] if (ri-4) < len(phase_colors) else "FFFFFF"
        cl(ws, ri, 1, item.get("week", ""),     bold=True, bg=bg, color="FFFFFF")
        cl(ws, ri, 2, item.get("focus", ""),    bold=True, bg="F7F9FC")
        cl(ws, ri, 3, item.get("actions", ""))
        cl(ws, ri, 4, item.get("milestone", ""), italic=True, color="375623", bg="E2EFDA")
        c = cl(ws, ri, 5, "⬜ Not Started", align="center", bg="FFF3EC")
        c.font = Font(name="Arial", size=10, color="ED7D31", bold=True)
        rh(ws, ri, 70)
    legend_row = 4 + len(data.get("action_plan", [])) + 1
    ws.merge_cells(f"A{legend_row}:E{legend_row}")
    cl(ws, legend_row, 1, "STATUS: ⬜ Not Started  |  🔄 In Progress  |  ✅ Done", italic=True, bg="EBF3FB", color="1F3864", size=9)
    rh(ws, legend_row, 18)
    cw(ws, 1, 16); cw(ws, 2, 26); cw(ws, 3, 44); cw(ws, 4, 34); cw(ws, 5, 18)


def build_b_fear_sheet(ws, data):
    section_header(ws, 1, "YOUR FEAR REFRAME", span=2)
    fear = data.get("fear_reframe", {})
    rows = [
        ("THE FEAR",           fear.get("fear", ""),     "C00000", "FFE0E0"),
        ("THE REAL TRUTH",     fear.get("truth", ""),    "375623", "E2EFDA"),
        ("ACTION THIS WEEK 1", fear.get("action_1", ""), "2E75B6", "EBF3FB"),
        ("ACTION THIS WEEK 2", fear.get("action_2", ""), "2E75B6", "EBF3FB"),
        ("ACTION THIS WEEK 3", fear.get("action_3", ""), "2E75B6", "EBF3FB"),
    ]
    for ri, (label, val, label_color, val_bg) in enumerate(rows, 3):
        cl(ws, ri, 1, label, bold=True, bg="D9E1F2", color=label_color, size=11)
        cl(ws, ri, 2, val, bg=val_bg, size=11)
        rh(ws, ri, 60)
    cw(ws, 1, 22); cw(ws, 2, 70)


def build_b_revenue_sheet(ws, data):
    section_header(ws, 1, "REVENUE PROJECTION", span=9)
    col_headers(ws, 3, ["PROGRAM", "PRICE", "Y1 UNITS (LOW)", "Y1 REV (LOW)", "Y1 UNITS (HIGH)", "Y1 REV (HIGH)", "Y2 UNITS", "Y2 REVENUE", "YoY GROWTH"])
    items = data.get("revenue", [])
    data_start = 4
    for ri, item in enumerate(items, data_start):
        bg = "F7F9FC" if ri % 2 == 0 else "FFFFFF"
        cl(ws, ri, 1, item.get("program", ""), bold=True, bg=bg, color="1F3864")
        c2 = cl(ws, ri, 2, item.get("price_numeric", 0), bg=bg, align="right"); money_fmt(c2)
        c3 = cl(ws, ri, 3, item.get("year1_low_units", 0),    bg=bg, align="center")
        c4 = cl(ws, ri, 4, item.get("year1_low_revenue", 0),  bg=bg, align="right"); money_fmt(c4)
        c5 = cl(ws, ri, 5, item.get("year1_high_units", 0),   bg=bg, align="center")
        c6 = cl(ws, ri, 6, item.get("year1_high_revenue", 0), bg=bg, align="right"); money_fmt(c6)
        c7 = cl(ws, ri, 7, item.get("year2_units", 0),        bg=bg, align="center")
        c8 = cl(ws, ri, 8, item.get("year2_revenue", 0),      bg=bg, align="right"); money_fmt(c8)
        c9 = ws.cell(row=ri, column=9, value=f"=IF(F{ri}=0,\"\",(H{ri}-F{ri})/F{ri})")
        c9.font  = Font(name="Arial", size=10, bold=True, color="375623")
        c9.fill  = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        c9.alignment = Alignment(horizontal="center", vertical="top")
        pct_fmt(c9); rh(ws, ri, 30)
    tr = data_start + len(items)
    cl(ws, tr, 1, "TOTAL", bold=True, bg="1F3864", color="FFFFFF")
    cl(ws, tr, 2, "", bg="1F3864"); cl(ws, tr, 3, "", bg="D9E1F2"); cl(ws, tr, 5, "", bg="D9E1F2"); cl(ws, tr, 7, "", bg="D9E1F2")
    for col_idx, col_letter in [(4, "D"), (6, "F"), (8, "H")]:
        c = ws.cell(row=tr, column=col_idx, value=f"=SUM({col_letter}{data_start}:{col_letter}{tr-1})")
        c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        c.alignment = Alignment(horizontal="right", vertical="top"); money_fmt(c)
    ct = ws.cell(row=tr, column=9, value=f"=IF(F{tr}=0,\"\",(H{tr}-F{tr})/F{tr})")
    ct.font = Font(name="Arial", size=11, bold=True, color="FFD700")
    ct.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    ct.alignment = Alignment(horizontal="center", vertical="top"); pct_fmt(ct)
    rh(ws, tr, 26)
    note_row = tr + 1
    ws.merge_cells(f"A{note_row}:I{note_row}")
    cl(ws, note_row, 1, data.get("revenue_note", ""), italic=True, bg="EBF3FB", color="1F3864")
    rh(ws, note_row, 45)
    cw(ws, 1, 34); cw(ws, 2, 12); cw(ws, 3, 14); cw(ws, 4, 17); cw(ws, 5, 14); cw(ws, 6, 17); cw(ws, 7, 12); cw(ws, 8, 17); cw(ws, 9, 14)


def build_b_trust_channels_sheet(ws, data):
    section_header(ws, 1, "TRUST CHANNEL BENCHMARKS", span=3)
    col_headers(ws, 3, ["METRIC", "PARTNERSHIPS", "REFERRALS"])
    bench = data.get("trust_channel_benchmarks", {})
    p = bench.get("partnerships", {}); r = bench.get("referrals", {})
    def yn(v): return "Yes" if str(v).strip().lower() == "yes" else "No"
    rows = [
        ("USED NOW?",                  yn(p.get("used_now","No")),  yn(r.get("used_now","No"))),
        ("WHY UNDERESTIMATED",         p.get("why_underestimated",""),   r.get("why_underestimated","")),
        ("EST. ANNUAL CLIENTS",        p.get("estimated_annual_clients",0), r.get("estimated_annual_clients",0)),
        ("EST. ANNUAL REVENUE",        p.get("estimated_annual_revenue",0), r.get("estimated_annual_revenue",0)),
        ("BENCHMARK NOTE",             p.get("benchmark_note",""),    r.get("benchmark_note","")),
        ("FIRST 3 ACTIONS",            "\n".join(f"• {a}" for a in p.get("first_actions",[]) if a), "\n".join(f"• {a}" for a in r.get("first_actions",[]) if a)),
    ]
    for ri, (label, p_val, r_val) in enumerate(rows, 4):
        bg = "F7F9FC" if ri % 2 == 0 else "FFFFFF"
        cl(ws, ri, 1, label, bold=True, bg="D9E1F2", color="1F3864")
        pc = cl(ws, ri, 2, p_val, bg=bg); rc_ = cl(ws, ri, 3, r_val, bg=bg)
        if "REVENUE" in label: money_fmt(pc); money_fmt(rc_)
        rh(ws, ri, 70 if "ACTIONS" in label else 55)
    cw(ws, 1, 32); cw(ws, 2, 46); cw(ws, 3, 46)


def build_b_competitor_sheet(ws, data):
    section_header(ws, 1, "COMPETITOR ANALYSIS", span=3)
    COMP_COLORS = [("4472C4","EBF3FB"), ("ED7D31","FFF3EC"), ("70AD47","EFF7E8")]
    FIELD_LABELS = [
        ("WEBSITE","url"), ("NICHE","niche"), ("STRATEGY","strategy"), ("CONTENT APPROACH","content_approach"),
        ("BUSINESS MODEL","business_model"), ("FLAGSHIP OFFER","flagship_offer"), ("FLAGSHIP PRICE","flagship_price"),
        ("FUNNEL STRUCTURE","funnel_structure"), ("EST. ANNUAL REVENUE","estimated_revenue"),
        ("AUDIENCE SIZE","audience_size"), ("STRENGTHS","strengths"), ("WEAKNESSES","weaknesses"), ("YOUR EDGE","your_edge"),
    ]
    current_row = 3
    for idx, comp in enumerate(data.get("competitors", [])[:3]):
        hbg, bbg = COMP_COLORS[idx]
        cl(ws, current_row, 1, f"COMPETITOR {idx+1} — {comp.get('name','')}", bold=True, color="FFFFFF", bg=hbg, size=11)
        ws.merge_cells(f"A{current_row}:C{current_row}"); rh(ws, current_row, 22); current_row += 1
        col_headers(ws, current_row, ["FIELD","DETAIL","YOUR NOTES"], bg=hbg); current_row += 1
        for label, key in FIELD_LABELS:
            value = comp.get(key, "")
            if key == "your_edge":
                cl(ws, current_row, 1, label, bold=True, bg="1F3864", color="FFD700", size=10)
                cl(ws, current_row, 2, value, bold=True, bg="1F3864", color="FFFFFF", size=10)
                cl(ws, current_row, 3, "", bg="1F3864")
            else:
                alt_bg = bbg if current_row % 2 == 0 else "FFFFFF"
                cl(ws, current_row, 1, label, bold=True, bg="D9E1F2", color="1F3864")
                cl(ws, current_row, 2, value, bg=alt_bg); cl(ws, current_row, 3, "", bg=alt_bg)
            rh(ws, current_row, 50); current_row += 1
        current_row += 1
    cw(ws, 1, 26); cw(ws, 2, 52); cw(ws, 3, 45)


def build_b_paid_sheet(ws, data):
    section_header(ws, 1, "PAID ACQUISITION — Ads to Webinar to Discovery Call", span=2)
    rows = [
        ("AD GOAL",           "Drive registrations to your free webinar. Sell a free training, not coaching."),
        ("AD HOOK FORMULA",   "Line 1: Call out the exact person + pain. Line 2: Name the real unspoken problem. Line 3: Introduce the webinar. Line 4: One CTA with genuine urgency."),
        ("WEBINAR — Self Recognition", "Describe their daily reality so precisely they feel you're reading their diary."),
        ("WEBINAR — Epiphany",         "Show them the problem isn't what they think. Introduce your framework. Max 3 insights."),
        ("WEBINAR — Invitation",       "Invite the RIGHT people to a call. Name who it's NOT for."),
        ("DISCOVERY CALL GOAL",        "Determine fit. Never convince. If you're selling, you've already lost."),
        ("CALL — Listen For",          "Specific answers. They take responsibility. Their goal matches your outcome."),
        ("CALL — Red Flags",           "First question is about price. Blames external factors. Expects results in two weeks."),
        ("CALL — Close",               "YES: Here's how we start. NOT YET: What needs to change. NO: Not the right fit. Never discount."),
    ]
    for ri, (label, val) in enumerate(rows, 3):
        cl(ws, ri, 1, label, bold=True, bg="D9E1F2", color="1F3864", size=10)
        cl(ws, ri, 2, val, bg="F7F9FC" if ri % 2 == 0 else "FFFFFF")
        rh(ws, ri, 55)
    cw(ws, 1, 28); cw(ws, 2, 75)


def build_b_upgrade_sheet(ws, payment_url: str = ""):
    section_header(ws, 1, "🔒 YOUR FULL PLAN — UNLOCK NOW", span=2)
    locked_sections = [
        ("😨 Fear Reframe (Full)",             "The real truth + 3 specific actions to take this week."),
        ("📣 Marketing Strategy (Full)",        "5 tailored channels, sample posts, 2-week calendar, KPIs."),
        ("💰 Revenue Projection (Complete)",    "Year 1 high, Year 2, YoY growth + personalised revenue note."),
        ("🤝 Trust Channel Benchmarks",         "Partnerships + Referrals with annual client/revenue benchmarks."),
        ("🕵 Competitor Analysis (Full)",       "All 3 competitors — pricing, funnel, weaknesses, your edge."),
        ("📈 Paid Acquisition Blueprint",       "Ad hook formula, webinar script, discovery call framework."),
        ("📅 90-Day Plan — Weeks 5–12",         "Convert → close → systematise → scale."),
        ("💼 Funnel Revenue — Mid + Flagship",  "Monthly projections for your top two revenue tiers."),
    ]
    for ri, (label, desc) in enumerate(locked_sections, 3):
        cl(ws, ri, 1, f"🔒  {label}", bold=True, bg="D9E1F2", color="1F3864", size=11)
        cl(ws, ri, 2, f"{desc}\n\n➡  Unlock in the Full Plan.", bg="FFF8E7", color="856404")
        rh(ws, ri, 70)
    cta_row = 3 + len(locked_sections) + 1
    ws.merge_cells(f"A{cta_row}:B{cta_row}")
    cl(ws, cta_row, 1,
       "✨  UPGRADE TO THE FULL PLAN  ✨\n\nYou've seen your Offer Blueprint, partial Funnel, and previews of every section.\n"
       "Join coaches who upgraded this month to unlock the complete strategy.",
       bold=True, bg="1F3864", color="FFFFFF", size=12)
    rh(ws, cta_row, 120)
    link_row = cta_row + 1
    ws.merge_cells(f"A{link_row}:B{link_row}")
    tag_cell = cl(ws, link_row, 1, "👉  CLICK HERE TO UNLOCK YOUR FULL PLAN  →",
                  bold=True, bg="FFD700", color="1F3864", size=14, align="center")
    if payment_url: tag_cell.hyperlink = payment_url
    rh(ws, link_row, 40)
    cw(ws, 1, 32); cw(ws, 2, 72)


# ─────────────────────────────────────────────────────────────────────────────
# PATH B — FREE (locked) VARIANTS
# ─────────────────────────────────────────────────────────────────────────────

def build_b_free_funnel_sheet(ws, data):
    section_header(ws, 1, "YOUR RECOMMENDED FUNNEL  ·  Preview", span=8)
    col_headers(ws, 3, ["TIER","NAME","FORMAT","PRICE","PURPOSE","EST. MONTHLY CLIENTS","MONTHLY REV (LOW)","MONTHLY REV (HIGH)"])
    tier_colors = {"LEAD MAGNET":"70AD47","LOW-TICKET":"4472C4","MID-TICKET":"ED7D31","FLAGSHIP":"C00000"}
    for ri, item in enumerate(data.get("funnel",[]), 4):
        tier = item.get("tier",""); bg = tier_colors.get(tier,"FFFFFF"); locked = tier in ("MID-TICKET","FLAGSHIP")
        cl(ws, ri, 1, f"🔒 {tier}" if locked else tier, bold=True, bg=LOCK_LABEL_BG if locked else bg, color=LOCK_LABEL_FG if locked else "FFFFFF")
        if locked:
            _lock(ws, ri, 2, "🔒 [Unlock Identity]"); _lock(ws, ri, 3, "🔒 [Unlock Structure]")
        else:
            cl(ws, ri, 2, item.get("name","")); cl(ws, ri, 3, item.get("format",""))
        cl(ws, ri, 4, item.get("price",""), bold=True, color="1F3864", bg=LOCK_BG if locked else "FFFFFF")
        if locked:
            cl(ws, ri, 5, _tease(item.get("purpose",""),50), italic=True, color=LOCK_FG, bg=LOCK_BG)
            _lock(ws, ri, 6, align="center"); _lock(ws, ri, 7, "🔒 Upgrade", align="right"); _lock(ws, ri, 8, "🔒 Upgrade", align="right")
        else:
            cl(ws, ri, 5, item.get("purpose",""))
            c6 = cl(ws, ri, 6, item.get("monthly_clients",0), align="center")
            c7 = cl(ws, ri, 7, item.get("monthly_revenue_low",0), align="right")
            c8 = cl(ws, ri, 8, item.get("monthly_revenue_high",0), align="right")
            money_fmt(c7); money_fmt(c8)
        rh(ws, ri, 52)
    tr = 4 + len(data.get("funnel",[]))
    cl(ws, tr, 1, "MONTHLY TOTALS", bold=True, bg="1F3864", color="FFFFFF")
    for c in range(2,6): cl(ws, tr, c, "", bg="D9E1F2")
    _lock(ws, tr, 6, align="center"); _lock(ws, tr, 7, "🔒 Full totals in paid plan", align="right"); _lock(ws, tr, 8, "🔒 Full totals in paid plan", align="right")
    rh(ws, tr, 22)
    cw(ws, 1, 16); cw(ws, 2, 26); cw(ws, 3, 22); cw(ws, 4, 14); cw(ws, 5, 40); cw(ws, 6, 20); cw(ws, 7, 18); cw(ws, 8, 18)


def build_b_free_action_sheet(ws, data):
    section_header(ws, 1, "YOUR 90-DAY ACTION PLAN  ·  Preview", span=5)
    col_headers(ws, 3, ["WEEK","FOCUS","ACTIONS","MILESTONE","STATUS"])
    phase_colors = ["4472C4","4472C4","ED7D31","ED7D31","C00000"]
    for ri, item in enumerate(data.get("action_plan",[]), 4):
        idx = ri-4; locked = idx >= 2; blackout = idx >= 3
        bg_ph = phase_colors[idx] if idx < len(phase_colors) else "FFFFFF"
        cl(ws, ri, 1, f"🔒 {item.get('week','')}" if locked else item.get("week",""), bold=True, bg=LOCK_LABEL_BG if locked else bg_ph, color=LOCK_LABEL_FG if locked else "FFFFFF")
        if blackout: _lock(ws, ri, 2, "🔒 [Strategy Locked]")
        else: cl(ws, ri, 2, item.get("focus",""), bold=True, bg="F2F2F2" if locked else "F7F9FC", color=LOCK_LABEL_FG if locked else "000000")
        if locked:
            cl(ws, ri, 3, _tease(item.get("actions",""), keep=40 if blackout else 60), italic=True, color=LOCK_FG, bg=LOCK_BG)
            cl(ws, ri, 4, _tease(item.get("milestone",""), keep=20 if blackout else 35), italic=True, color=LOCK_FG, bg=LOCK_BG)
            _lock(ws, ri, 5, align="center")
        else:
            cl(ws, ri, 3, item.get("actions",""))
            cl(ws, ri, 4, item.get("milestone",""), italic=True, color="375623", bg="E2EFDA")
            c = cl(ws, ri, 5, "⬜ Not Started", align="center", bg="FFF3EC")
            c.font = Font(name="Arial", size=10, color="ED7D31", bold=True)
        rh(ws, ri, 70)
    legend_row = 4 + len(data.get("action_plan",[])) + 1
    ws.merge_cells(f"A{legend_row}:E{legend_row}")
    cl(ws, legend_row, 1, "Weeks 1–4 fully unlocked.  🔒 Weeks 5–12 in the Full Plan.", italic=True, bg="EBF3FB", color="1F3864", size=9)
    rh(ws, legend_row, 18)
    cw(ws, 1, 16); cw(ws, 2, 28); cw(ws, 3, 44); cw(ws, 4, 34); cw(ws, 5, 18)


def build_b_free_fear_sheet(ws, data):
    section_header(ws, 1, "YOUR FEAR REFRAME  ·  Preview", span=2)
    fear = data.get("fear_reframe",{})
    rows = [
        ("THE FEAR",           fear.get("fear",""),     "C00000","FFE0E0", False),
        ("THE REAL TRUTH",     fear.get("truth",""),    "375623","E2EFDA", True),
        ("ACTION THIS WEEK 1", fear.get("action_1",""), "2E75B6","EBF3FB", True),
        ("ACTION THIS WEEK 2", fear.get("action_2",""), "2E75B6","EBF3FB", True),
        ("ACTION THIS WEEK 3", fear.get("action_3",""), "2E75B6","EBF3FB", True),
    ]
    for ri, (label, val, label_color, val_bg, locked) in enumerate(rows, 3):
        cl(ws, ri, 1, label, bold=True, bg="D9E1F2", color=label_color, size=11)
        if locked: cl(ws, ri, 2, _tease(val, keep=55), italic=True, color=LOCK_FG, bg=LOCK_BG, size=11)
        else: cl(ws, ri, 2, val, bg=val_bg, size=11)
        rh(ws, ri, 60)
    cw(ws, 1, 22); cw(ws, 2, 70)


def build_b_free_revenue_sheet(ws, data):
    section_header(ws, 1, "REVENUE PROJECTION  ·  Preview", span=9)
    col_headers(ws, 3, ["PROGRAM","PRICE","Y1 UNITS (LOW)","Y1 REVENUE (LOW)","Y1 UNITS (HIGH)","Y1 REVENUE (HIGH)","Y2 UNITS","Y2 REVENUE","YoY GROWTH"])
    items = data.get("revenue",[]); data_start = 4
    for ri, item in enumerate(items, data_start):
        bg = "F7F9FC" if ri % 2 == 0 else "FFFFFF"
        cl(ws, ri, 1, item.get("program",""), bold=True, bg=bg, color="1F3864")
        c2 = cl(ws, ri, 2, item.get("price_numeric",0), bg=bg, align="right"); money_fmt(c2)
        if ri > data_start:
            for col in range(3,10): _lock(ws, ri, col, "🔒")
        else:
            cl(ws, ri, 3, item.get("year1_low_units",0), bg=bg, align="center")
            c4 = cl(ws, ri, 4, item.get("year1_low_revenue",0), bg=bg, align="right"); money_fmt(c4)
            for col in range(5,10): _lock(ws, ri, col, "🔒")
        rh(ws, ri, 30)
    tr = data_start + len(items)
    cl(ws, tr, 1, "TOTAL", bold=True, bg="1F3864", color="FFFFFF"); cl(ws, tr, 2, "", bg="1F3864")
    for col in range(3,10): _lock(ws, tr, col, "🔒 Full plan only")
    rh(ws, tr, 26)
    note_row = tr+1; ws.merge_cells(f"A{note_row}:I{note_row}")
    cl(ws, note_row, 1, "🔒 Full Year 1 (high), Year 2 projections in the Full Plan.", italic=True, bg="EBF3FB", color="1F3864")
    rh(ws, note_row, 40)
    cw(ws, 1, 34); cw(ws, 2, 12); cw(ws, 3, 14); cw(ws, 4, 17); cw(ws, 5, 14); cw(ws, 6, 17); cw(ws, 7, 12); cw(ws, 8, 17); cw(ws, 9, 14)


def build_b_free_trust_channels_sheet(ws, data):
    section_header(ws, 1, "TRUST CHANNEL BENCHMARKS  ·  Preview", span=3)
    col_headers(ws, 3, ["METRIC","PARTNERSHIPS","REFERRALS"])
    bench = data.get("trust_channel_benchmarks",{})
    p = bench.get("partnerships",{}); r = bench.get("referrals",{})
    def yn(v): return "Yes" if str(v).strip().lower()=="yes" else "No"
    p_used = yn(p.get("used_now","No")); r_used = yn(r.get("used_now","No"))
    rows = [
        ("USED NOW?",           p_used, r_used, False),
        ("WHY UNDERESTIMATED",  p.get("why_underestimated",""), r.get("why_underestimated",""), False),
        ("EST. ANNUAL CLIENTS", p.get("estimated_annual_clients",0), r.get("estimated_annual_clients",0), True),
        ("EST. ANNUAL REVENUE", p.get("estimated_annual_revenue",0), r.get("estimated_annual_revenue",0), True),
        ("BENCHMARK NOTE",      p.get("benchmark_note",""), r.get("benchmark_note",""), True),
        ("FIRST 3 ACTIONS",     "\n".join(f"• {a}" for a in p.get("first_actions",[]) if a), "\n".join(f"• {a}" for a in r.get("first_actions",[]) if a), True),
    ]
    for ri, (label, p_val, r_val, lockable) in enumerate(rows, 4):
        bg = "F7F9FC" if ri % 2 == 0 else "FFFFFF"
        cl(ws, ri, 1, label, bold=True, bg="D9E1F2", color="1F3864")
        def _wc(col, val, used):
            if lockable and used == "No": return _lock(ws, ri, col, "🔒 [Unlock in Full Plan]")
            return cl(ws, ri, col, val, bg=bg)
        pc = _wc(2, p_val, p_used); rc_ = _wc(3, r_val, r_used)
        if "REVENUE" in label:
            if pc: money_fmt(pc)
            if rc_: money_fmt(rc_)
        rh(ws, ri, 70 if "ACTIONS" in label else 55)
    cw(ws, 1, 32); cw(ws, 2, 46); cw(ws, 3, 46)


def build_b_free_competitor_sheet(ws, data):
    section_header(ws, 1, "COMPETITOR ANALYSIS  ·  Preview", span=3)
    COMP_COLORS = [("4472C4","EBF3FB"),("ED7D31","FFF3EC"),("70AD47","EFF7E8")]
    VISIBLE = {"url","niche","strengths","your_edge"}
    FIELD_LABELS = [("WEBSITE","url"),("NICHE","niche"),("STRATEGY","strategy"),("CONTENT APPROACH","content_approach"),
                    ("BUSINESS MODEL","business_model"),("FLAGSHIP OFFER","flagship_offer"),("FLAGSHIP PRICE","flagship_price"),
                    ("FUNNEL STRUCTURE","funnel_structure"),("EST. ANNUAL REVENUE","estimated_revenue"),("AUDIENCE SIZE","audience_size"),
                    ("STRENGTHS","strengths"),("WEAKNESSES","weaknesses"),("YOUR EDGE","your_edge")]
    current_row = 3
    for idx, comp in enumerate(data.get("competitors",[])[:3]):
        hbg, bbg = COMP_COLORS[idx]; is_secret = idx > 0
        label_text = f"🔒 COMPETITOR {idx+1} — [Unlock Identity]" if is_secret else f"COMPETITOR {idx+1} — {comp.get('name','')}"
        cl(ws, current_row, 1, label_text, bold=True, color="FFFFFF", bg=hbg if not is_secret else LOCK_FG, size=11)
        ws.merge_cells(f"A{current_row}:C{current_row}"); rh(ws, current_row, 22); current_row += 1
        col_headers(ws, current_row, ["FIELD","DETAIL","YOUR NOTES"], bg=hbg if not is_secret else LOCK_FG); current_row += 1
        for label, key in FIELD_LABELS:
            value = comp.get(key,""); locked = (key not in VISIBLE) or is_secret
            alt_bg = bbg if current_row % 2 == 0 else "FFFFFF"
            if key == "your_edge" and not is_secret:
                cl(ws, current_row, 1, label, bold=True, bg="1F3864", color="FFD700", size=10)
                cl(ws, current_row, 2, value, bold=True, bg="1F3864", color="FFFFFF", size=10)
                cl(ws, current_row, 3, "", bg="1F3864")
            elif locked:
                cl(ws, current_row, 1, label, bold=True, bg="D9E1F2", color="1F3864")
                cl(ws, current_row, 2, _tease(value, keep=30 if is_secret else 50), italic=True, color=LOCK_FG, bg=LOCK_BG)
                _lock(ws, current_row, 3, "🔒 [Unlock Details]")
            else:
                cl(ws, current_row, 1, label, bold=True, bg="D9E1F2", color="1F3864")
                cl(ws, current_row, 2, value, bg=alt_bg); cl(ws, current_row, 3, "", bg=alt_bg)
            rh(ws, current_row, 50); current_row += 1
        current_row += 1
    cw(ws, 1, 26); cw(ws, 2, 52); cw(ws, 3, 45)


def build_b_free_paid_sheet(ws, data):
    section_header(ws, 1, "PAID ACQUISITION  ·  Preview", span=2)
    rows_config = [
        ("AD GOAL", True), ("AD HOOK FORMULA", False), ("WEBINAR — Self Recognition", False),
        ("WEBINAR — Epiphany", False), ("WEBINAR — Invitation", False),
        ("DISCOVERY CALL GOAL", True), ("CALL — Listen For", False), ("CALL — Red Flags", False), ("CALL — Close", False),
    ]
    rows_content = [
        "Drive registrations to your free webinar. You are selling a free training, not coaching.",
        "Line 1: Call out the exact person and their pain. Line 2: Name the real problem. Line 3: Introduce the webinar. Line 4: One CTA with urgency.",
        "Describe their daily reality so precisely they feel you are reading their diary.",
        "Show them the problem is not what they think. Your framework is the real explanation.",
        "Invite the RIGHT people. Name who it is NOT for. Frame it as a fit conversation.",
        "Determine fit. Never convince. The right client should be selling themselves.",
        "Specific answers. They take responsibility. Their goal matches your outcome.",
        "First question is about price. Blames external factors. Expects results in two weeks.",
        "YES: Here is how we start. NOT YET: What needs to change. NO: Not the right fit. Never discount.",
    ]
    for ri, ((label, visible), val) in enumerate(zip(rows_config, rows_content), 3):
        cl(ws, ri, 1, label, bold=True, bg="D9E1F2", color="1F3864", size=10)
        if visible: cl(ws, ri, 2, val, bg="F7F9FC" if ri%2==0 else "FFFFFF")
        else: cl(ws, ri, 2, _tease(val, keep=60), italic=True, color=LOCK_FG, bg=LOCK_BG)
        rh(ws, ri, 55)
    cw(ws, 1, 28); cw(ws, 2, 75)


# ─────────────────────────────────────────────────────────────────────────────
# PATH B MARKETING
# ─────────────────────────────────────────────────────────────────────────────

CHANNEL_PALETTES = [
    {"header": "0A66C2", "sub": "1A85D6", "light": "D0E8FF", "accent": "EBF5FF"},
    {"header": "C13584", "sub": "D63C8A", "light": "FFD6EC", "accent": "FFF0F7"},
    {"header": "217346", "sub": "2A8F58", "light": "C6EFCE", "accent": "F0FFF4"},
    {"header": "CC0000", "sub": "E00000", "light": "FFD0D0", "accent": "FFF5F5"},
    {"header": "7C3AED", "sub": "8B5CF6", "light": "EDE9FE", "accent": "F9F7FF"},
]
CHANNEL_EMOJIS = {
    "LinkedIn": "💼", "Instagram": "📸", "Email": "📧", "Email Newsletter": "📧",
    "YouTube": "🎬", "Podcast": "🎙", "Facebook Ads": "📣", "Blog": "✍️",
    "Blog/SEO": "✍️", "Twitter": "🐦", "Twitter/X": "🐦", "Webinar": "📡",
    "Partnerships": "🤝", "Referrals": "🔁",
}

AUDIT_MAP = {
    "headline_before": ("Headline", "headline_after"),
    "about_before": ("About / Summary", "about_after"),
    "banner_tip": ("Banner Image", None),
    "featured_section": ("Featured Section", None),
    "bio_before": ("Bio", "bio_after"),
    "bio_formula": ("Bio Formula", None),
    "highlight_covers": ("Story Highlights", None),
    "story_strategy": ("Stories Strategy", None),
    "subject_line_formula": ("Subject Line Formula", None),
    "welcome_sequence": ("Welcome Sequence", None),
    "list_growth_tactic": ("List Growth Tactic", None),
    "segmentation_tip": ("Segmentation", None),
    "channel_description_rewrite": ("Channel Description", None),
    "about_page_cta": ("About Page CTA", None),
    "thumbnail_formula": ("Thumbnail Formula", None),
    "channel_trailer_script_outline": ("Trailer Script", None),
    "show_description_rewrite": ("Show Description", None),
    "episode_title_formula": ("Episode Title Formula", None),
    "intro_hook_script": ("Intro Hook Script", None),
    "guest_pitch_template": ("Guest Pitch Template", None),
    "audience_definition": ("Target Audience", None),
    "ad_hook_formula": ("Ad Hook Formula", None),
    "creative_brief": ("Creative Brief", None),
    "landing_page_checklist": ("Landing Page", None),
    "partner_target_list": ("Partner Target List", None),
    "partner_offer": ("Partner Offer", None),
    "outreach_script": ("Outreach Script", None),
    "co_marketing_asset": ("Co-Marketing Asset", None),
    "referral_offer": ("Referral Offer", None),
    "ask_points": ("Ask Points", None),
    "referral_script": ("Referral Script", None),
    "tracking_system": ("Tracking System", None),
}


def build_all_marketing_sheets(wb, data, is_free=False, payment_url=""):
    for idx, ch in enumerate(data.get("marketing", [])):
        palette = CHANNEL_PALETTES[idx % len(CHANNEL_PALETTES)]
        ch_name = ch.get("channel", f"Channel {idx+1}")
        emoji   = CHANNEL_EMOJIS.get(ch_name, "📌")
        ws      = wb.create_sheet(f"{emoji} {ch_name}"[:31])
        _build_channel_sheet(ws, ch, palette, is_free, payment_url)


def _build_channel_sheet(ws, ch, palette, is_free, payment_url):
    H = palette["header"]; SB = palette["sub"]; LT = palette["light"]; AC = palette["accent"]
    name = ch.get("channel","Channel"); priority = ch.get("priority","Primary")
    why = ch.get("why_this_channel",""); audit = ch.get("profile_audit",{})
    pillars = ch.get("content_pillars",[]); calendar = ch.get("two_week_calendar",[])
    qwins = ch.get("quick_wins",[]); kpis = ch.get("kpis",{})
    row = 1

    ws.merge_cells(f"A{row}:F{row}")
    cl(ws, row, 1, f"{name.upper()} STRATEGY  ·  {priority} Channel", bold=True, color="FFFFFF", bg=H, size=14, align="center")
    rh(ws, row, 32); row += 1
    ws.merge_cells(f"A{row}:F{row}")
    cl(ws, row, 1, why, italic=True, color="1F3864", bg=AC, size=10)
    rh(ws, row, 28); row += 2

    # Profile Audit
    ws.merge_cells(f"A{row}:F{row}")
    cl(ws, row, 1, "① PROFILE AUDIT — Fix These First", bold=True, color="FFFFFF", bg=H, size=11)
    rh(ws, row, 22); row += 1
    cl(ws, row, 1, "FIELD", bold=True, color="FFFFFF", bg=SB, size=9, align="center")
    cl(ws, row, 2, "❌  BEFORE", bold=True, color="FFFFFF", bg="C00000", size=9)
    cl(ws, row, 3, "YOUR CURRENT ✏️", bold=True, color="555555", bg="F2F2F2", size=9, italic=True)
    ws.merge_cells(f"D{row}:F{row}")
    cl(ws, row, 4, "✅  AFTER — YOUR REWRITE", bold=True, color="FFFFFF", bg="375623", size=9)
    rh(ws, row, 18); row += 1

    rendered = set()
    for key, (label, after_key) in AUDIT_MAP.items():
        val = audit.get(key)
        if not val or key in rendered: continue
        rendered.add(key)
        if after_key: rendered.add(after_key)
        after_val = audit.get(after_key, "") if after_key else ""
        bg_row = AC if row % 2 == 0 else "FFFFFF"
        cl(ws, row, 1, label, bold=True, bg=LT, color="1F3864", size=9)
        if after_key:
            cl(ws, row, 2, val, italic=True, color="9E2A2B", bg="FFE8E8", size=9)
            cl(ws, row, 3, "", bg="F2F2F2", size=9)
            ws.merge_cells(f"D{row}:F{row}")
            cl(ws, row, 4, after_val, bold=True, color="1F4E24", bg="E2EFDA", size=9)
        else:
            ws.merge_cells(f"B{row}:F{row}")
            cl(ws, row, 2, val, bg=bg_row, size=9)
        rh(ws, row, max(40, min(130, len(str(val or "")+str(after_val or "")) // 3)))
        row += 1
    row += 1

    # Content Pillars
    ws.merge_cells(f"A{row}:F{row}")
    cl(ws, row, 1, "② CONTENT PILLARS", bold=True, color="FFFFFF", bg=H, size=11)
    rh(ws, row, 22); row += 1
    for hdr, col_n in [("PILLAR",1),("PURPOSE",2),("FORMAT & FREQ",3)]:
        cl(ws, row, col_n, hdr, bold=True, color="FFFFFF", bg=SB, size=9)
    ws.merge_cells(f"D{row}:F{row}")
    cl(ws, row, 4, "3 SCROLL-STOPPING HOOKS", bold=True, color="FFFFFF", bg=SB, size=9)
    rh(ws, row, 18); row += 1
    for pi, pillar in enumerate(pillars):
        bg_p = AC if pi % 2 == 0 else "FFFFFF"
        cl(ws, row, 1, pillar.get("pillar",""),  bold=True, bg=LT, color="1F3864", size=10)
        cl(ws, row, 2, pillar.get("purpose",""), bg=bg_p, size=9)
        cl(ws, row, 3, f"{pillar.get('formats','')}  ·  {pillar.get('frequency','')}", bg=bg_p, size=9, italic=True)
        ws.merge_cells(f"D{row}:F{row}")
        cl(ws, row, 4, "\n".join(f"▸  {h}" for h in pillar.get("hooks",[])), bg=bg_p, size=9, color="1F3864")
        rh(ws, row, 75); row += 1
    row += 1

    # Sample Posts
    ws.merge_cells(f"A{row}:F{row}")
    cl(ws, row, 1, "③ SAMPLE POSTS" + ("  🔒  Full posts in Full Plan" if is_free else " — Copy, Edit, Post"), bold=True, color="FFFFFF", bg=H, size=11)
    rh(ws, row, 22); row += 1
    for pi, pillar in enumerate(pillars):
        post = pillar.get("sample_post",""); bg_p = AC if pi % 2 == 0 else "FFFFFF"
        cl(ws, row, 1, pillar.get("pillar",f"Pillar {pi+1}"), bold=True, bg=LT, color="1F3864", size=9)
        ws.merge_cells(f"B{row}:F{row}")
        if is_free:
            cl(ws, row, 2, (post[:90].rstrip()+"…  🔒 Upgrade") if len(post)>90 else post+"  🔒", italic=True, color="9E9E9E", bg="F2F2F2", size=9)
            rh(ws, row, 40)
        else:
            cl(ws, row, 2, post, bg=bg_p, size=9)
            rh(ws, row, max(80, min(200, len(post)//2)))
        row += 1
    row += 1

    # 2-Week Calendar
    ws.merge_cells(f"A{row}:F{row}")
    cl(ws, row, 1, "④ 2-WEEK CONTENT CALENDAR" + ("  🔒  Week 2 in Full Plan" if is_free else " — Copy, Paste, Post"), bold=True, color="FFFFFF", bg=H, size=11)
    rh(ws, row, 22); row += 1
    for hdr, col_n in [("DAY",1),("FORMAT",2),("PILLAR",3)]:
        cl(ws, row, col_n, hdr, bold=True, color="FFFFFF", bg=SB, size=9, align="center")
    ws.merge_cells(f"D{row}:E{row}")
    cl(ws, row, 4, "HOOK / OPENING LINE", bold=True, color="FFFFFF", bg=SB, size=9)
    cl(ws, row, 6, "CTA", bold=True, color="FFFFFF", bg=SB, size=9, align="center")
    rh(ws, row, 18); row += 1
    for ci, entry in enumerate(calendar):
        bg_c = AC if ci % 2 == 0 else "FFFFFF"
        locked = is_free and ci >= 5
        if locked:
            cl(ws, row, 1, entry.get("day",""), bold=True, bg="D9D9D9", color="9E9E9E", size=9, align="center")
            ws.merge_cells(f"B{row}:F{row}")
            cl(ws, row, 2, "🔒 Week 2 in Full Plan", italic=True, color="9E9E9E", bg="F2F2F2", size=9)
        else:
            cl(ws, row, 1, entry.get("day",""),    bold=True, bg=LT, color="1F3864", size=9, align="center")
            cl(ws, row, 2, entry.get("format",""), bg=bg_c, size=9, align="center")
            cl(ws, row, 3, entry.get("pillar",""), bg=bg_c, size=9)
            ws.merge_cells(f"D{row}:E{row}")
            cl(ws, row, 4, entry.get("hook",""), bg=bg_c, size=9)
            cl(ws, row, 6, entry.get("cta",""),  bg="E2EFDA", color="375623", size=9, bold=True)
        rh(ws, row, 45); row += 1
    row += 1

    # Quick Wins
    ws.merge_cells(f"A{row}:F{row}")
    cl(ws, row, 1, "⑤ QUICK WINS" + ("  🔒  Unlock in Full Plan" if is_free else " — Do These in 48 Hours"), bold=True, color="FFFFFF", bg=H, size=11)
    rh(ws, row, 22); row += 1
    for qi, win in enumerate(qwins, 1):
        bg_q = AC if qi % 2 == 0 else "FFFFFF"
        cl(ws, row, 1, f"ACTION {qi}", bold=True, bg=LT, color="1F3864", size=10, align="center")
        ws.merge_cells(f"B{row}:E{row}")
        if is_free:
            cl(ws, row, 2, (win[:80].rstrip()+"…  🔒") if len(win)>80 else win+"  🔒", italic=True, color="9E9E9E", bg="F2F2F2", size=9)
        else:
            cl(ws, row, 2, win, bg=bg_q, size=9)
        cl(ws, row, 6, "⬜ Done", bg="FFF3EC", color="ED7D31", bold=True, size=10, align="center")
        rh(ws, row, max(45, min(100, len(win)//2))); row += 1
    row += 1

    # KPIs
    ws.merge_cells(f"A{row}:F{row}")
    cl(ws, row, 1, "⑥ SUCCESS METRICS", bold=True, color="FFFFFF", bg=H, size=11)
    rh(ws, row, 22); row += 1
    kpi_fields = [("Posting Frequency","posting_frequency"),("Engagement Rate","engagement_rate_target"),
                  ("Connection Growth","connection_growth_per_month"),("Leads per Month","leads_per_month"),("🎯 Primary Metric","primary_metric")]
    for ki, (label, key) in enumerate(kpi_fields):
        bg_k = AC if ki % 2 == 0 else "FFFFFF"; val = kpis.get(key,"—"); is_primary = key=="primary_metric"
        cl(ws, row, 1, label, bold=True, bg=LT, color="1F3864", size=9)
        ws.merge_cells(f"B{row}:C{row}")
        cl(ws, row, 2, val, bold=is_primary, bg="E2EFDA" if is_primary else bg_k, color="375623" if is_primary else "000000", size=10)
        ws.merge_cells(f"D{row}:E{row}")
        cl(ws, row, 4, "YOUR CURRENT →", italic=True, color="9E9E9E", bg="F2F2F2", size=8, align="center")
        cl(ws, row, 6, "⬜ Track", italic=True, color="9E9E9E", bg="F2F2F2", size=8, align="center")
        rh(ws, row, max(28, min(80, len(str(val))//2))); row += 1

    cw(ws, 1, 22); cw(ws, 2, 30); cw(ws, 3, 22); cw(ws, 4, 32); cw(ws, 5, 18); cw(ws, 6, 20)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN BUILD FUNCTIONS — branching on coach_path A / B / C
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(data: dict, file_path: str):
    """Full (paid) report — all three paths."""
    wb   = Workbook()
    path = data.get("coach_path", "B")

    if path in ("C", "D"):
        ws1 = wb.active; ws1.title = "🔍 Hidden Ceiling"
        build_c_diagnosis_sheet(ws1, data, is_free=False)
        wb.create_sheet("🚀 Expansion Offer");      build_c_offer_sheet(wb["🚀 Expansion Offer"], data, is_free=False)
        wb.create_sheet("📊 Revenue Scenarios");    build_c_revenue_comparison_sheet(wb["📊 Revenue Scenarios"], data, is_free=False)
        wb.create_sheet("📣 Marketing New Offer");  build_c_marketing_sheet(wb["📣 Marketing New Offer"], data, is_free=False)
        wb.create_sheet("📅 90-Day Plan");          build_c_action_sheet(wb["📅 90-Day Plan"], data, is_free=False)
        wb.create_sheet("🕵 Competitors");          build_c_competitor_sheet(wb["🕵 Competitors"], data, is_free=False)
        wb.create_sheet("😨 Fear Reframe");         build_c_fear_sheet(wb["😨 Fear Reframe"], data, is_free=False)

    elif path == "A":
        ws1 = wb.active; ws1.title = "📊 Revenue Ceiling"
        build_a_revenue_ceiling_sheet(ws1, data, is_free=False)
        ws2 = wb.create_sheet("🏢 Corporate Offer")
        build_a_corporate_offer_sheet(ws2, data, is_free=False)
        ws3 = wb.create_sheet("🎤 Speaking Strategy")
        build_a_speaking_sheet(ws3, data, is_free=False)
        ws4 = wb.create_sheet("📅 90-Day Expansion")
        build_a_action_sheet(ws4, data, is_free=False)
        ws5 = wb.create_sheet("😨 Fear Reframe")
        build_a_fear_sheet(ws5, data, is_free=False)

    else:  # Path B
        ws1 = wb.active; ws1.title = "Your Offer"
        build_b_offer_sheet(ws1, data, is_premium=True)
        wb.create_sheet("Your Funnel");          build_b_funnel_sheet(wb["Your Funnel"], data)
        wb.create_sheet("90-Day Plan");          build_b_action_sheet(wb["90-Day Plan"], data)
        wb.create_sheet("Fear Reframe");         build_b_fear_sheet(wb["Fear Reframe"], data)
        build_all_marketing_sheets(wb, data, is_free=False)
        wb.create_sheet("Revenue");              build_b_revenue_sheet(wb["Revenue"], data)
        wb.create_sheet("Trust Channels");       build_b_trust_channels_sheet(wb["Trust Channels"], data)
        wb.create_sheet("Competitor Analysis");  build_b_competitor_sheet(wb["Competitor Analysis"], data)
        if data.get("include_paid_funnel"):
            wb.create_sheet("Paid Acquisition"); build_b_paid_sheet(wb["Paid Acquisition"], data)

    wb.save(file_path)


def build_free_excel(data: dict, file_path: str, payment_url: str = ""):
    """Free teaser — all three paths with locked sections."""
    wb   = Workbook()
    path = data.get("coach_path", "B")

    if path in ("C", "D"):
        ws1 = wb.active; ws1.title = "🔍 Hidden Ceiling"
        build_c_diagnosis_sheet(ws1, data, is_free=True)
        wb.create_sheet("🚀 Expansion Offer");      build_c_offer_sheet(wb["🚀 Expansion Offer"], data, is_free=True)
        wb.create_sheet("📊 Revenue Scenarios");    build_c_revenue_comparison_sheet(wb["📊 Revenue Scenarios"], data, is_free=True)
        wb.create_sheet("📣 Marketing New Offer");  build_c_marketing_sheet(wb["📣 Marketing New Offer"], data, is_free=True)
        wb.create_sheet("📅 90-Day Plan");          build_c_action_sheet(wb["📅 90-Day Plan"], data, is_free=True)
        wb.create_sheet("🕵 Competitors");          build_c_competitor_sheet(wb["🕵 Competitors"], data, is_free=True)
        wb.create_sheet("😨 Fear Reframe");         build_c_fear_sheet(wb["😨 Fear Reframe"], data, is_free=True)
        ws_cta = wb.create_sheet("🔒 Unlock Full Plan")
        build_c_upgrade_sheet(ws_cta, payment_url)

    elif path == "A":
        ws1 = wb.active; ws1.title = "📊 Revenue Ceiling"
        build_a_revenue_ceiling_sheet(ws1, data, is_free=True)
        ws2 = wb.create_sheet("🏢 Corporate Offer")
        build_a_corporate_offer_sheet(ws2, data, is_free=True)
        ws3 = wb.create_sheet("🎤 Speaking Strategy")
        build_a_speaking_sheet(ws3, data, is_free=True)
        ws4 = wb.create_sheet("📅 90-Day Expansion")
        build_a_action_sheet(ws4, data, is_free=True)
        ws5 = wb.create_sheet("😨 Fear Reframe")
        build_a_fear_sheet(ws5, data, is_free=True)
        ws_cta = wb.create_sheet("🔒 Unlock Full Plan")
        build_a_upgrade_sheet(ws_cta, payment_url)

    else:  # Path B
        ws1 = wb.active; ws1.title = "Your Offer"
        build_b_offer_sheet(ws1, data, is_premium=False)
        ws2 = wb.create_sheet("Your Funnel");         build_b_free_funnel_sheet(ws2, data)
        ws3 = wb.create_sheet("90-Day Plan");         build_b_free_action_sheet(ws3, data)
        ws4 = wb.create_sheet("Fear Reframe");        build_b_free_fear_sheet(ws4, data)
        build_all_marketing_sheets(wb, data, is_free=True, payment_url=payment_url)
        ws6 = wb.create_sheet("Revenue");             build_b_free_revenue_sheet(ws6, data)
        ws7 = wb.create_sheet("Trust Channels");      build_b_free_trust_channels_sheet(ws7, data)
        ws8 = wb.create_sheet("Competitor Analysis"); build_b_free_competitor_sheet(ws8, data)
        if data.get("include_paid_funnel"):
            ws9 = wb.create_sheet("Paid Acquisition"); build_b_free_paid_sheet(ws9, data)
        ws_cta = wb.create_sheet("🔒 Unlock Full Plan")
        build_b_upgrade_sheet(ws_cta, payment_url)

    wb.save(file_path)


# ─────────────────────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/webhook/tally", methods=["POST"])
def receive_tally():
    data   = request.json
    fields = data["data"]["fields"]
    answers = {}
    respondent_email = ""
    respondent_name  = ""

    for field in fields:
        label      = field.get("label","").strip()
        value      = field.get("value","")
        field_type = field.get("type","")
        if isinstance(value, list):
            value = value[0] if len(value)==1 else ", ".join(str(v) for v in value)
        str_value = str(value) if value else ""
        answers[label] = str_value
        if field_type=="INPUT_EMAIL" or "email" in label.lower():
            respondent_email = str_value
        if not respondent_name and any(k in label.lower() for k in ["name","first name","full name"]):
            respondent_name = str_value

    file_id      = str(uuid.uuid4())[:8]
    plan         = call_openai(answers)
    _PLAN_CACHE[file_id] = plan

    free_path    = f"/tmp/free_{file_id}.xlsx"
    premium_path = f"/tmp/plan_{file_id}.xlsx"
    payment_url  = f"{request.host_url}payment?file_id={file_id}"
    build_free_excel(plan, free_path, payment_url)
    build_excel(plan, premium_path)

    if respondent_email:
        schedule_email_sequence(respondent_email, respondent_name or "Coach", file_id, request.host_url)

    return jsonify({
        "success": True,
        "free_download_url": f"{request.host_url}download/free/{file_id}",
        "premium_url": payment_url,
        "file_id": file_id,
        "coach_path": plan.get("coach_path","B"),
    })


@app.route("/submit-quiz", methods=["POST"])
def submit_quiz():
    form_data        = request.form
    respondent_email = form_data.get("email","")
    respondent_name  = form_data.get("name","")

    if not respondent_email:
        return "Email is required to receive your plan.", 400

    # Build answers dict — quiz uses Q1-Q17 field names
    answers = {
        "name":  respondent_name,
        "email": respondent_email,
        # Map quiz form fields to answer keys used by detection + prompts
        "Q1":  form_data.get("Q1", ""),   # Expertise
        "Q2":  form_data.get("Q2", ""),   # Natural advice
        "Q3":  form_data.get("Q3", ""),   # Experience level  ← KEY for Path C detection
        "Q4":  form_data.get("Q4", ""),   # Ideal client      ← KEY
        "Q5":  form_data.get("Q5", ""),   # Pain point        ← KEY
        "Q6":  form_data.get("Q6", ""),   # Desire
        "Q7":  form_data.get("Q7", ""),   # Exclusion
        "Q8":  form_data.get("Q8", ""),   # Delivery
        "Q9":  form_data.get("Q9", ""),   # Duration
        "Q10": form_data.get("Q10", ""),  # Pricing           ← KEY
        "Q12": form_data.get("Q12", ""),  # Audience size
        "Q13": form_data.get("Q13", ""),  # Biggest fear
        "Q14": form_data.get("Q14", ""),  # Hours/week
        "Q16": form_data.get("Q16", ""),  # Vision
        "Q15_interest": form_data.get("Q15_interest", "Yes"),
        "Q15": form_data.get("Q15", ""),  # Ad budget
    }

    # Multi-select channels
    channel_vals = [v for v in form_data.getlist("Q17") if v.strip()]
    other_channel = form_data.get("Q17_other", "").strip()
    if other_channel: channel_vals.append(other_channel)
    answers["Q17"] = ", ".join(channel_vals)   # ← KEY for Path C detection

    # Legacy Path A fork support
    qa_fork = form_data.get("QA_FORK", "").lower()
    answers["QA_FORK"] = qa_fork
    if qa_fork in ("a", "fully_booked", "path_a") or "fully" in qa_fork:
        for key in ["QB1","QB2","QB3","QB4","QB5","QB6","QB7","QB8"]:
            answers[key] = form_data.get(key, "")

    file_id      = str(uuid.uuid4())[:8]
    plan         = call_openai(answers)
    _PLAN_CACHE[file_id] = plan

    payment_url  = f"{request.host_url}payment?file_id={file_id}"
    free_path    = f"/tmp/free_{file_id}.xlsx"
    premium_path = f"/tmp/plan_{file_id}.xlsx"
    build_free_excel(plan, free_path, payment_url)
    build_excel(plan, premium_path)

    unsub_token = str(uuid.uuid4())
    _UNSUB_TOKENS[unsub_token] = {"email": respondent_email, "file_id": file_id}
    unsubscribe_url = f"{request.host_url}unsubscribe/{unsub_token}"
    html_email_1 = email_1_html(
        respondent_name or "Coach",
        f"{request.host_url}download/free/{file_id}",
        payment_url,
        unsubscribe_url,
    )
    send_email(respondent_email, "Your free Coaching Business Snapshot is here 🎯", html_email_1, free_path)

    try: scheduler.remove_job(f"abandoned_{respondent_email}")
    except: pass

    schedule_email_sequence(respondent_email, respondent_name or "Coach", file_id, request.host_url)
    return redirect(f"/payment?file_id={file_id}")


@app.route("/track-start", methods=["POST"])
def track_start():
    data  = request.get_json(silent=True) or {}
    email = data.get("email"); name = data.get("name")
    if not email: return jsonify({"success": False, "reason": "No email"}), 400
    unsub_token = str(uuid.uuid4())
    _UNSUB_TOKENS[unsub_token] = {"email": email}
    unsubscribe_url = f"{request.host_url}unsubscribe/{unsub_token}"
    quiz_url = f"{request.host_url}quiz"
    html = email_abandoned_html(name, quiz_url, unsubscribe_url)
    scheduler.add_job(send_email, trigger="date", run_date=datetime.now()+timedelta(minutes=30),
                      args=[email, "Did you get interrupted?", html], id=f"abandoned_{email}", replace_existing=True)
    return jsonify({"success": True})


@app.route("/unsubscribe/<token>")
def unsubscribe(token):
    data = _UNSUB_TOKENS.get(token)
    if not data: return "<h1>Invalid or expired link.</h1>", 400
    file_id = data.get("file_id"); email = data["email"]
    for delay in [0, 1, 2, 4, 6]:
        try: scheduler.remove_job(f"{file_id}_email_{delay}")
        except: pass
    try: scheduler.remove_job(f"abandoned_{email}")
    except: pass
    if token in _UNSUB_TOKENS: del _UNSUB_TOKENS[token]
    return f"<h1>Unsubscribed.</h1><p>No more emails to {email}.</p>"


@app.route("/test-email-sequence", methods=["POST"])
def test_email_sequence():
    body     = request.get_json(silent=True) or {}
    to_email = body.get("email","")
    name     = body.get("name","Test Coach")
    if not to_email: return jsonify({"error": "email is required"}), 400
    file_id     = "testtest"; host_url = request.host_url
    free_url    = f"{host_url}download/free/{file_id}"
    payment_url = f"{host_url}payment?file_id={file_id}"
    unsubscribe_url = f"{host_url}unsubscribe/test-token"
    html = email_1_html(name, free_url, payment_url, unsubscribe_url)
    send_email(to_email, "[TEST] Your free Coaching Business Snapshot is here 🎯", html)
    return jsonify({"success": True, "message": f"Test Email 1 sent to {to_email}"})


@app.route("/")
def index(): return render_template("landing.html")

@app.route("/quiz")
def quiz_page(): return render_template("quiz.html")

@app.route("/health")
def health():
    return jsonify({
        "status": "ok",
        "paths_supported": ["A", "B", "C"],
        "path_c_description": "Established 1:1 coach with untapped corporate/speaking/group potential"
    })


@app.route("/download/free/<file_id>")
def download_free_file(file_id):
    if not file_id.isalnum(): return "Invalid file ID", 400
    file_path = f"/tmp/free_{file_id}.xlsx"
    if not os.path.exists(file_path): return "File not found or expired", 404
    return send_file(file_path, as_attachment=True, download_name="Your_Free_Coaching_Snapshot.xlsx")


@app.route("/download/<file_id>")
def download_file(file_id):
    if not file_id.isalnum(): return "Invalid file ID", 400
    file_path = f"/tmp/plan_{file_id}.xlsx"
    if not os.path.exists(file_path): return "File not found or expired", 404
    return send_file(file_path, as_attachment=True, download_name="Your_Full_Coaching_Business_Plan.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# STRIPE
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/payment', methods=['GET'])
def payment():
    key     = os.environ.get('STRIPE_PUBLISHABLE_KEY')
    file_id = request.args.get('file_id','')
    return render_template('payment.html', key=key, file_id=file_id)


@app.route('/create-checkout-session', methods=['POST'])
def create_checkout_session():
    if not stripe.api_key:
        return jsonify(error="STRIPE_SECRET_KEY not set"), 500
    try:
        body    = request.get_json(silent=True) or {}
        file_id = body.get('file_id','')
        tier    = body.get('tier','standard')
        if tier == 'premium':
            price_amount = 60000; product_name = 'Full Plan + 1-on-1 Video Review'
        else:
            price_amount = 15000; product_name = 'Full Coaching Business Plan'
        session = stripe.checkout.Session.create(
            payment_method_types=['card'],
            line_items=[{'price_data': {'currency': 'usd', 'product_data': {'name': product_name}, 'unit_amount': price_amount}, 'quantity': 1}],
            mode='payment',
            success_url=request.host_url + f'success?file_id={file_id}&tier={tier}',
            cancel_url=request.host_url + 'cancel',
            metadata={'file_id': file_id, 'tier': tier},
        )
        return jsonify({'id': session.id})
    except stripe.error.StripeError as e:
        return jsonify(error=f"Stripe error: {e.user_message or str(e)}"), 400
    except Exception as e:
        app.logger.error(f"Checkout error: {e}")
        return jsonify(error="Internal error"), 500


@app.route('/success')
def success():
    file_id  = request.args.get('file_id','')
    tier     = request.args.get('tier','standard')
    tier_msg = ("Your full Coaching Business Plan is ready."
                if tier != 'premium' else
                "Your full Plan is ready, and we'll be in touch shortly for your 1-on-1 Video Review!")
    if file_id and file_id.isalnum():
        download_url = f"{request.host_url}download/{file_id}"
        return f"""
        <html><body style="font-family:Arial;text-align:center;padding:60px;">
          <h1 style="color:#4338ca;">🎉 Payment confirmed!</h1>
          <p style="font-size:18px;">{tier_msg}</p>
          <a href="{download_url}" style="display:inline-block;margin-top:20px;background:#4338ca;color:white;padding:14px 32px;border-radius:8px;text-decoration:none;font-weight:bold;font-size:16px;">
            ⬇ Download Your Full Plan
          </a>
        </body></html>"""
    return "<h1>Thanks for your order!</h1><p>Check your email for your report.</p>"


@app.route('/cancel')
def cancel(): return "<h1>Order cancelled.</h1><p><a href='/'>Return home</a></p>"


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=True, host="0.0.0.0", port=port)